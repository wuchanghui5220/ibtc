import math
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict




from openpyxl import Workbook
import re

def generate_device_names_excel(device_names, filename):
    wb = Workbook()
    
    # Create sheets
    sheets = {
        'GPU': wb.active,
        'Leaf': wb.create_sheet(title="Leaf"),
        'Spine': wb.create_sheet(title="Spine"),
        'Core': wb.create_sheet(title="Core")
    }
    sheets['GPU'].title = "GPU"

    # Define the headers
    headers = ['hostname', 'new hostname', 'locations', 'U位']

    def sort_key(name):
        # Extract numbers from the name
        numbers = re.findall(r'\d+', name)
        # Convert extracted numbers to integers
        return [int(num) for num in numbers]

    # Write headers and device names to respective sheets
    for device_type, names in device_names.items():
        sheet = sheets[device_type]
        # Write the headers first
        sheet.append(headers)
        # Then write the sorted device names
        for name in sorted(names, key=sort_key):
            # Assuming the name is the 'Source Device', leaving other fields blank for now
            row = [name] + [''] * (len(headers) - 1)
            sheet.append(row)

    # Save the workbook
    wb.save(filename)
    print(f"Device names have been written to {filename}")


def generate_flexible_topology_256(num_servers, ports_per_switch=64):
    # ... (previous code remains the same)
        # 计算需要的SU数量和Leaf交换机数量
    num_sus = math.ceil(num_servers / 32)
    num_leafs = num_sus * 8
    
    # 计算Spine交换机数量（必须是32的因数）
    spine_factors = [2, 4, 8, 16, 32]
    num_spines = next(factor for factor in spine_factors if factor >= num_leafs / 2)

    # 初始化连接列表
    leaf_spine_connections = []
    server_leaf_connections = []

    # 计算每个Leaf到Spine的连接数
    connections_per_spine = 32 // num_spines

    # 生成Leaf到Spine的连接
    for leaf_index in range(num_leafs):
        su_index = leaf_index // 8
        local_leaf_index = leaf_index % 8
        leaf_name = f"LG{su_index+1}-Leaf{local_leaf_index+1}"
        
        for spine_index in range(num_spines):
            spine_name = f"Spine{spine_index+1}"
            for port_offset in range(connections_per_spine):
                leaf_port = spine_index * connections_per_spine + port_offset + 1
                spine_port = leaf_index * connections_per_spine + port_offset + 1
                leaf_spine_connections.append((leaf_name, f"{leaf_port}", spine_name, f"{spine_port}"))

    # 生成服务器到Leaf的连接
    for server_index in range(num_servers):
        su_index = server_index // 32
        local_server_index = server_index % 32
        gpu_index = local_server_index + 1

        for hca_index in range(1, 9):  # 每个GPU有8个HCA
            leaf_index = hca_index - 1
            leaf_port = 33 + local_server_index

            server_name = f"SU{su_index+1}-GPU{gpu_index}"
            hca_port = f"mlx5_{hca_index}"
            leaf_name = f"LG{su_index+1}-Leaf{leaf_index+1}"
            server_leaf_connections.append((server_name, hca_port, leaf_name, f"{leaf_port}"))

    device_names = {
        'GPU': set(),
        'Leaf': set(),
        'Spine': set(),
        'Core': set()
    }

    # Collect device names
    for leaf_index in range(num_leafs):
        su_index = leaf_index // 8
        local_leaf_index = leaf_index % 8
        leaf_name = f"LG{su_index+1}-Leaf{local_leaf_index+1}"
        device_names['Leaf'].add(leaf_name)
    
    for spine_index in range(num_spines):
        spine_name = f"Spine{spine_index+1}"
        device_names['Spine'].add(spine_name)
    
    for server_index in range(num_servers):
        su_index = server_index // 32
        local_server_index = server_index % 32
        gpu_index = local_server_index + 1
        server_name = f"SU{su_index+1}-GPU{gpu_index}"
        device_names['GPU'].add(server_name)

    return leaf_spine_connections + server_leaf_connections, device_names

def generate_three_tier_topology_512(total_servers, num_cgs=8, leaves_per_lg=8, spines_per_sg=8, cores_per_cg=8, ports_per_leaf=64, ports_per_spine=64):
    # ... (previous code remains the same)
    # 根据服务器总数计算LG和SG的数量
    servers_per_lg = 32  # 假设每个LG有32台服务器
    num_lgs = math.ceil(total_servers / servers_per_lg)
    num_sgs = num_lgs  # SG的数量通常与LG相同

    # 初始化连接列表
    connections = []

    # 生成Leaf到Spine的连接
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            for spine_index in range(spines_per_sg):
                leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
                spine_name = f"SG{lg_index+1}-Spine{spine_index+1}"
                for port_group in range(32 // spines_per_sg):
                    leaf_port = spine_index * (32 // spines_per_sg) + port_group + 1
                    spine_port = leaf_index * (32 // spines_per_sg) + port_group + 1
                    connections.append((leaf_name, f"{leaf_port}", spine_name, f"{spine_port}"))

    # 生成Spine到Core的连接
    for sg_index in range(num_sgs):
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{sg_index+1}-Spine{spine_index+1}"
            # 为每个Spine交换机确定连接到Core交换机的端口范围
            spine_ports = list(range(33, 65))
            for core_index in range(cores_per_cg):
                core_name = f"CG{spine_index+1}-Core{core_index+1}"
                # 为每个Core交换机确定连接到Spine交换机的端口范围
                core_ports = list(range(1 + sg_index * 4, 5 + sg_index * 4))
                for spine_port, core_port in zip(spine_ports[core_index*4:(core_index+1)*4], core_ports):
                    connections.append((spine_name, f"{spine_port}", core_name, f"{core_port}"))                
                # for spine_port, core_port in zip(spine_ports, core_ports):
                #     connections.append((spine_name, f"{spine_port}", core_name, f"{core_port}"))

    # 生成GPU到Leaf的连接
    gpus_per_lg = 32
    hcas_per_gpu = 8
    for lg_index in range(num_lgs):
        # 如果是最后一个LG，确定剩余的GPU数量
        gpus_in_this_lg = gpus_per_lg if (lg_index + 1) < num_lgs else total_servers - lg_index * gpus_per_lg
        for gpu_index in range(gpus_in_this_lg):
            for hca_index in range(hcas_per_gpu):
                su_name = f"SU{lg_index+1}"
                gpu_name = f"{su_name}-GPU{gpu_index+1}"
                hca_name = f"{gpu_name}"
                hca_port = f"mlx5_{hca_index}"
                leaf_name = f"LG{lg_index+1}-Leaf{hca_index+1}"
                leaf_port = 33 + gpu_index
                connections.append((hca_name, hca_port, leaf_name, f"{leaf_port}"))
    
    device_names = {
        'GPU': set(),
        'Leaf': set(),
        'Spine': set(),
        'Core': set()
    }

    # Collect device names
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
            device_names['Leaf'].add(leaf_name)
        
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{lg_index+1}-Spine{spine_index+1}"
            device_names['Spine'].add(spine_name)
        
        for gpu_index in range(min(gpus_per_lg, total_servers - lg_index * gpus_per_lg)):
            gpu_name = f"SU{lg_index+1}-GPU{gpu_index+1}"
            device_names['GPU'].add(gpu_name)

    for cg_index in range(num_cgs):
        for core_index in range(cores_per_cg):
            core_name = f"CG{cg_index+1}-Core{core_index+1}"
            device_names['Core'].add(core_name)

    return connections, device_names

def generate_three_tier_topology_1024(total_servers, num_cgs=8, leaves_per_lg=8, spines_per_sg=8, cores_per_cg=16, ports_per_leaf=64, ports_per_spine=64):
    # Calculate the number of LGs and SGs based on the total number of servers
    servers_per_lg = 32  # Assuming each LG has 32 servers
    num_lgs = math.ceil(total_servers / servers_per_lg)
    num_sgs = num_lgs  # SGs number is usually the same as LGs

    # Initialize connection list
    connections = []

    # Generate Leaf to Spine connections
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            for spine_index in range(spines_per_sg):
                leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
                # The spine group index is based on the modulus of the LG index
                spine_name = f"SG{(lg_index % num_sgs) + 1}-Spine{spine_index+1}"
                for port_group in range(4):  # 32 ports divided into 8 groups, 4 cables per group
                    leaf_port = spine_index * 4 + port_group + 1
                    spine_port = leaf_index * 4 + port_group + 1
                    connections.append((leaf_name, f"{leaf_port}", spine_name, f"{spine_port}"))

    # Generate Spine to Core connections
    for sg_index in range(num_sgs):
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{sg_index+1}-Spine{spine_index+1}"
            spine_ports = list(range(33, 65))  # Using the second half of ports (33-64)
            for core_index in range(cores_per_cg):
                core_name = f"CG{spine_index+1}-Core{core_index+1}"
                core_ports = list(range(sg_index * 2 + 1, sg_index * 2 + 3))  # 2 cables per connection
                for spine_port, core_port in zip(spine_ports[core_index*2:(core_index+1)*2], core_ports):
                    connections.append((spine_name, f"{spine_port}", core_name, f"{core_port}"))

    # Generate GPU to Leaf connections
    gpus_per_lg = 32
    hcas_per_gpu = 8
    for lg_index in range(num_lgs):
        # Determine the number of GPUs for this LG, which could be less for the last LG
        gpus_in_this_lg = gpus_per_lg if (lg_index + 1) < num_lgs else total_servers - lg_index * gpus_per_lg
        for gpu_index in range(gpus_in_this_lg):
            for hca_index in range(hcas_per_gpu):
                su_name = f"SU{lg_index+1}"
                gpu_name = f"{su_name}-GPU{gpu_index+1}"
                hca_name = f"{gpu_name}"
                hca_port = f"mlx5_{hca_index}"
                leaf_name = f"LG{lg_index+1}-Leaf{hca_index+1}"
                leaf_port = 33 + gpu_index
                connections.append((hca_name, hca_port, leaf_name, f"{leaf_port}"))    
    device_names = {
        'GPU': set(),
        'Leaf': set(),
        'Spine': set(),
        'Core': set()
    }

    # Collect device names
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
            device_names['Leaf'].add(leaf_name)
        
        spine_group = (lg_index % num_sgs) + 1
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{spine_group}-Spine{spine_index+1}"
            device_names['Spine'].add(spine_name)
        
        for gpu_index in range(min(gpus_per_lg, total_servers - lg_index * gpus_per_lg)):
            gpu_name = f"SU{lg_index+1}-GPU{gpu_index+1}"
            device_names['GPU'].add(gpu_name)

    for cg_index in range(num_cgs):
        for core_index in range(cores_per_cg):
            core_name = f"CG{cg_index+1}-Core{core_index+1}"
            device_names['Core'].add(core_name)

    return connections, device_names

def generate_three_tier_topology_2048(total_servers, num_cgs=8, leaves_per_lg=8, spines_per_sg=8, cores_per_cg=32, ports_per_leaf=64, ports_per_spine=64):
    # Calculate the number of LGs and SGs based on the total number of servers
    servers_per_lg = 32  # Assuming each LG has 32 servers
    num_lgs = math.ceil(total_servers / servers_per_lg)
    num_sgs = num_lgs  # SGs number is usually the same as LGs

    # Initialize connection list
    connections = []

    # Generate Leaf to Spine connections
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            for spine_index in range(spines_per_sg):
                leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
                spine_name = f"SG{(lg_index % num_sgs)+1}-Spine{spine_index+1}"
                for port_group in range(4):  # 32 ports divided into 8 groups, 4 cables per group
                    leaf_port = spine_index * 4 + port_group + 1
                    spine_port = leaf_index * 4 + port_group + 1
                    connections.append((leaf_name, f"{leaf_port}", spine_name, f"{spine_port}"))

    # Generate Spine to Core connections
    for sg_index in range(num_sgs):
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{sg_index+1}-Spine{spine_index+1}"
            spine_ports = list(range(33, 65))  # Using the second half of ports (33-64)
            for core_index in range(cores_per_cg):
                cg_index = spine_index  # CG index corresponds to Spine index
                core_name = f"CG{cg_index+1}-Core{core_index+1}"
                core_ports = list(range(sg_index + 1, sg_index + 2))  # 1 cable per connection
                for spine_port, core_port in zip(spine_ports[core_index:core_index+1], core_ports):
                    connections.append((spine_name, f"{spine_port}", core_name, f"{core_port}"))

    # Generate GPU to Leaf connections
    gpus_per_lg = 32
    hcas_per_gpu = 8
    for lg_index in range(num_lgs):
        # Determine the number of GPUs for this LG, which could be less for the last LG
        gpus_in_this_lg = min(gpus_per_lg, total_servers - lg_index * gpus_per_lg)
        for gpu_index in range(gpus_in_this_lg):
            for hca_index in range(hcas_per_gpu):
                su_name = f"SU{lg_index+1}"
                gpu_name = f"{su_name}-GPU{gpu_index+1}"
                hca_name = f"{gpu_name}"
                hca_port = f"mlx5_{hca_index}"
                leaf_name = f"LG{lg_index+1}-Leaf{hca_index+1}"
                leaf_port = 33 + gpu_index
                connections.append((hca_name, hca_port, leaf_name, f"{leaf_port}"))
    device_names = {
        'GPU': set(),
        'Leaf': set(),
        'Spine': set(),
        'Core': set()
    }

    # Collect device names
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
            device_names['Leaf'].add(leaf_name)
        
        spine_group = (lg_index % num_sgs) + 1
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{spine_group}-Spine{spine_index+1}"
            device_names['Spine'].add(spine_name)
        
        for gpu_index in range(min(gpus_per_lg, total_servers - lg_index * gpus_per_lg)):
            gpu_name = f"SU{lg_index+1}-GPU{gpu_index+1}"
            device_names['GPU'].add(gpu_name)

    for cg_index in range(num_cgs):
        for core_index in range(cores_per_cg):
            core_name = f"CG{cg_index+1}-Core{core_index+1}"
            device_names['Core'].add(core_name)

    return connections, device_names

def generate_topology(total_servers):
    if 0 <= total_servers <= 256:
        return generate_flexible_topology_256(total_servers)
    elif 257 <= total_servers <= 512:
        return generate_three_tier_topology_512(total_servers)
    elif 513 <= total_servers <= 1024:
        return generate_three_tier_topology_1024(total_servers)
    elif 1025 <= total_servers <= 2048:
        return generate_three_tier_topology_2048(total_servers)
    else:
        print("Only the range between 1 and 2048 is supported.")
        exit()




from openpyxl import Workbook

def port_to_interface(port):
    if port.startswith('mlx5_'):
        return '1'
    port_num = int(port.replace('mlx5_', ''))
    return f"{(port_num - 1) // 2 + 1}/{(port_num - 1) % 2 + 1}"

def write_to_excel(connections, filename):
    # Create a new workbook
    wb = Workbook()
    
    # Create six sheets
    sheets = {
        'GPU to Leaf': wb.active,
        'Leaf to GPU': wb.create_sheet(title="Leaf to GPU"),
        'Leaf to Spine': wb.create_sheet(title="Leaf to Spine"),
        'Spine to Leaf': wb.create_sheet(title="Spine to Leaf"),
        'Spine to Core': wb.create_sheet(title="Spine to Core"),
        'Core to Spine': wb.create_sheet(title="Core to Spine")
    }
    sheets['GPU to Leaf'].title = "GPU to Leaf"

    # Write headers
    headers = ['Source Device', 'Source Port', 'Source Interface', 
               'Destination Device', 'Destination Port', 'Destination Interface']
    for sheet in sheets.values():
        sheet.append(headers)

    # Write connections to appropriate sheets
    for conn in connections:
        source, source_port, dest, dest_port = conn
        source_interface = port_to_interface(source_port)
        dest_interface = port_to_interface(dest_port)
        
        row = [source, source_port, source_interface, dest, dest_port, dest_interface]
        rev_row = [dest, dest_port, dest_interface, source, source_port, source_interface]
        
        if "GPU" in source and "Leaf" in dest:
            sheets['GPU to Leaf'].append(row)
            sheets['Leaf to GPU'].append(rev_row)
        elif "Leaf" in source and "Spine" in dest:
            sheets['Leaf to Spine'].append(row)
            sheets['Spine to Leaf'].append(rev_row)
        elif "Spine" in source and "Core" in dest:
            sheets['Spine to Core'].append(row)
            sheets['Core to Spine'].append(rev_row)

    # Save the workbook
    wb.save(filename)


from openpyxl import Workbook
from openpyxl.styles import Font



def natural_sort_key(s):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]

def generate_switch_port_mappings(connections, filename):
    # Create a new workbook
    wb = Workbook()
    
    # Create three sheets
    sheets = {
        'Leaf': wb.active,
        'Spine': wb.create_sheet(title="Spine"),
        'Core': wb.create_sheet(title="Core")
    }
    sheets['Leaf'].title = "Leaf"

    # Initialize dictionaries to store port mappings for each switch
    switch_port_mappings = {
        'Leaf': {},
        'Spine': {},
        'Core': {}
    }

    # Process connections to create port mappings
    for conn in connections:
        source, source_port, dest, dest_port = conn
        
        # Function to update port mappings
        def update_mapping(switch_type, switch, port, connected_to, connected_port):
            if switch not in switch_port_mappings[switch_type]:
                switch_port_mappings[switch_type][switch] = [''] * 64
            port_index = int(port.replace('Port', '')) - 1
            switch_port_mappings[switch_type][switch][port_index] = (connected_to, connected_port)

        # Update mappings for both source and destination
        for switch_type in ['Leaf', 'Spine', 'Core']:
            if switch_type in source:
                update_mapping(switch_type, source, source_port, dest, dest_port)
            if switch_type in dest:
                update_mapping(switch_type, dest, dest_port, source, source_port)

    # Write port mappings to sheets
    headers = ['Port', 'Interface', 'Connected Device', 'Connected Port', 'Connected Interface']
    for switch_type, mappings in switch_port_mappings.items():
        sheet = sheets[switch_type]
        sheet.append(headers)
        for switch in sorted(mappings.keys(), key=natural_sort_key):
            ports = mappings[switch]
            # Write switch name
            sheet.append([switch])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
            
            # Write port mappings
            for i, mapping in enumerate(ports, start=1):
                if mapping:
                    connected_device, connected_port = mapping
                    row = [f"{i}", port_to_interface(f"{i}"), 
                           connected_device, connected_port, port_to_interface(connected_port)]
                else:
                    row = [f"{i}", port_to_interface(f"{i}"), '', '', '']
                sheet.append(row)
            
            # Add an empty row for separation
            sheet.append([])

    # Save the workbook
    wb.save(filename)


def main():
    total_servers = int(input("Enter the total number of servers: "))
    connections, device_names = generate_topology(total_servers)
    
    # Print connections to screen
    print("Connections:")
    for conn in connections:
        print(f"{conn[0]}:{conn[1]} -> {conn[2]}:{conn[3]}")
    
    # Write connections to Excel file
    connections_filename = f"topology_{total_servers}_servers.xlsx"
    write_to_excel(connections, connections_filename)
    print(f"\nConnections have been written to {connections_filename}")

    # Generate new Excel file with device names
    device_names_filename = f"device_names_{total_servers}_servers.xlsx"
    generate_device_names_excel(device_names, device_names_filename)
    print(f"\nDevice names have been written to {device_names_filename}")

    # Generate new Excel file with switch port mappings
    switch_port_mappings_filename = f"switch_port_mappings_{total_servers}_servers.xlsx"
    generate_switch_port_mappings(connections, switch_port_mappings_filename)
    print(f"\nSwitch port mappings have been written to {switch_port_mappings_filename}")

if __name__ == "__main__":
    main()
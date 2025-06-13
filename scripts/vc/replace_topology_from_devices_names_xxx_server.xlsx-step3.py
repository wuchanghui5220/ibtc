import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import os
import time

# def read_device_mapping(filename):
#     """
#     Read device mapping from A.xlsx, considering multiple sheets
#     Returns a dictionary with hostnames as keys and dictionaries of corresponding data as values
#     """
#     try:
#         workbook = openpyxl.load_workbook(filename, read_only=True)
#         mapping = {}
#         for sheet in workbook.sheetnames:
#             ws = workbook[sheet]
#             print(f"Processing sheet: {sheet}")
#             headers = None
#             for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
#                 if row_index == 1:
#                     headers = row
#                     print(f"Headers found in {sheet}: {headers}")
#                     continue
#                 if len(row) >= 4 and row[0]:  # Ensure at least four columns and non-empty first column
#                     hostname = str(row[0])
#                     mapping[hostname] = {
#                         "new hostname": str(row[1]),
#                         "locations": str(row[2]),
#                         "U位": str(row[3])
#                     }
        
#         print(f"Total unique hostnames found: {len(mapping)}")
#         return mapping
#     except InvalidFileException:
#         print(f"Error: Unable to read {filename}. Make sure it's a valid Excel file.")
#     except Exception as e:
#         print(f"Unexpected error reading {filename}: {e}")
#     return None
def read_device_mapping(filename):
    """
    Read device mapping from A.xlsx, considering multiple sheets
    Returns a dictionary with hostnames as keys and dictionaries of corresponding data as values
    """
    try:
        workbook = openpyxl.load_workbook(filename, read_only=True)
        mapping = {}
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            print(f"Processing sheet: {sheet}")
            headers = None
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_index == 1:
                    headers = row
                    print(f"Headers found in {sheet}: {headers}")
                    continue
                if len(row) >= 7 and row[0]:  # 确保至少有7列（包括机房和坐标列）
                    hostname = str(row[0])
                    mapping[hostname] = {
                        "new hostname": str(row[1]),
                        "locations": str(row[2]),
                        "U位": str(row[3]),
                        "机房": str(row[4]) if row[4] is not None else "",  # E列
                        "设备坐标X": str(row[5]) if row[5] is not None else "",  # F列
                        "设备坐标Y": str(row[6]) if row[6] is not None else ""   # G列
                    }
        
        print(f"Total unique hostnames found: {len(mapping)}")
        return mapping
    except InvalidFileException:
        print(f"Error: Unable to read {filename}. Make sure it's a valid Excel file.")
    except Exception as e:
        print(f"Unexpected error reading {filename}: {e}")
    return None

# def update_topology_file(input_filename, output_filename, mapping):
    """
    Update the topology file with the mapping information
    """
    try:
        start_time = time.time()
        workbook = openpyxl.load_workbook(input_filename)
        updates_made = 0
        total_cells = sum(ws.max_row * ws.max_column for ws in workbook.worksheets)
        processed_cells = 0

        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            new_ws = workbook.create_sheet(title=f"{sheet}_updated")
            
            # Define new headers
            new_headers = [
                "Source Device", "Source locations", "Source U位", "Source Port", "Source Interface",
                "Destination Device", "Destination locations", "Destination U位", "Destination Port", "Destination Interface"
            ]
            new_ws.append(new_headers)

            for row in ws.iter_rows(min_row=2, values_only=True):
                source_device = str(row[0])
                dest_device = str(row[3])
                new_row = []

                # Process Source Device
                if source_device in mapping:
                    new_row.extend([
                        mapping[source_device]["new hostname"],
                        mapping[source_device]["locations"],
                        mapping[source_device]["U位"],

                    ])
                    updates_made += 1
                else:
                    new_row.extend([source_device, "", ""])

                # Add Source Port and Interface
                new_row.extend(row[1:3])

                # Process Destination Device
                if dest_device in mapping:
                    new_row.extend([
                        mapping[dest_device]["new hostname"],
                        mapping[dest_device]["locations"],
                        mapping[dest_device]["U位"],

                    ])
                    updates_made += 1
                else:
                    new_row.extend([dest_device, "", ""])

                # Add Destination Port and Interface
                new_row.extend(row[4:6])

                new_ws.append(new_row)

                processed_cells += len(row)
                if new_ws.max_row % 100 == 0:
                    progress = (processed_cells / total_cells) * 100
                    elapsed_time = time.time() - start_time
                    print(f"Progress: {progress:.2f}% - Rows processed: {new_ws.max_row} - Time elapsed: {elapsed_time:.2f}s", end='\r')

            workbook.remove(workbook[sheet])

        workbook.save(output_filename)
        print(f"\nTotal updates made: {updates_made}")
        print(f"Total time taken: {time.time() - start_time:.2f} seconds")
        return True
    except InvalidFileException:
        print(f"Error: Unable to read {input_filename}. Make sure it's a valid Excel file.")
    except PermissionError:
        print(f"Error: Unable to save {output_filename}. Make sure you have write permissions.")
    except Exception as e:
        print(f"Unexpected error processing {input_filename}: {e}")
        import traceback
        traceback.print_exc()
    return False
def update_topology_file(input_filename, output_filename, mapping):
    """
    Update the topology file with the mapping information
    """
    try:
        start_time = time.time()
        workbook = openpyxl.load_workbook(input_filename)
        updates_made = 0
        total_cells = sum(ws.max_row * ws.max_column for ws in workbook.worksheets)
        processed_cells = 0
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            new_ws = workbook.create_sheet(title=f"{sheet}_updated")
            
            # Define new headers
            new_headers = [
                "Source Device", "Source locations", "Source U位", "源机房", "源设备坐标X", "源设备坐标Y", "Source Port", "Source Interface",
                "Destination Device", "Destination locations", "Destination U位", "目标机房", "目标设备坐标X", "目标设备坐标Y", "Destination Port", "Destination Interface"
            ]
            new_ws.append(new_headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                source_device = str(row[0])
                dest_device = str(row[3])
                new_row = []
                # Process Source Device
                if source_device in mapping:
                    new_row.extend([
                        mapping[source_device].get("new hostname", ""),
                        mapping[source_device].get("locations", ""),
                        mapping[source_device].get("U位", ""),
                        mapping[source_device].get("机房", ""),
                        mapping[source_device].get("设备坐标X", ""),
                        mapping[source_device].get("设备坐标Y", "")
                    ])
                    updates_made += 1
                else:
                    # 为所有列提供空值
                    new_row.extend([""] * 6)
                # Add Source Port and Interface
                new_row.extend(row[1:3])
                # Process Destination Device
                if dest_device in mapping:
                    new_row.extend([
                        mapping[dest_device].get("new hostname", ""),
                        mapping[dest_device].get("locations", ""),
                        mapping[dest_device].get("U位", ""),
                        mapping[dest_device].get("机房", ""),
                        mapping[dest_device].get("设备坐标X", ""),
                        mapping[dest_device].get("设备坐标Y", "")
                    ])
                    updates_made += 1
                else:
                    # 为所有列提供空值
                    new_row.extend([""] * 6)
                # Add Destination Port and Interface
                new_row.extend(row[4:6])
                new_ws.append(new_row)
                processed_cells += len(row)
                if new_ws.max_row % 100 == 0:
                    progress = (processed_cells / total_cells) * 100
                    elapsed_time = time.time() - start_time
                    print(f"Progress: {progress:.2f}% - Rows processed: {new_ws.max_row} - Time elapsed: {elapsed_time:.2f}s", end='\r')
            workbook.remove(workbook[sheet])
        workbook.save(output_filename)
        print(f"\nTotal updates made: {updates_made}")
        print(f"Total time taken: {time.time() - start_time:.2f} seconds")
        return True
    except InvalidFileException:
        print(f"Error: Unable to read {input_filename}. Make sure it's a valid Excel file.")
    except PermissionError:
        print(f"Error: Unable to save {output_filename}. Make sure you have write permissions.")
    except Exception as e:
        print(f"Unexpected error processing {input_filename}: {e}")
        import traceback
        traceback.print_exc()
    return False


def main():
    mapping_file = 'device_names_256_servers_0529_v2.xlsx'
    input_file = 'topology_256_servers.xlsx'
    output_file = 'topology_256_servers-updated.xlsx'
    
    print("Reading device mapping from", mapping_file)
    mapping = read_device_mapping(mapping_file)
    
    if mapping is None or len(mapping) == 0:
        print(f"Error: No valid mappings found. Please check the {mapping_file} file.")
        return
    
    print(f"Updating topology data from {input_file} and saving to {output_file}")
    if update_topology_file(input_file, output_file, mapping):
        print("Update complete. Updated file saved as", output_file)
    else:
        print("Failed to complete the update process.")

if __name__ == "__main__":
    main()
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from typing import List, Dict, Tuple, Optional
import os
from datetime import datetime

def parse_cabinet_info(cell_value: str) -> Tuple[str, Tuple[int, int]]:
    """
    解析机柜信息，提取机柜编号和坐标
    例如: "25KW 机柜(A02)坐标(4,2)" -> ("A02", (4, 2))
    """
    # 提取机柜编号
    cabinet_match = re.search(r'机柜\s*[（(]([A-Z]\d+)[）)]', cell_value)
    cabinet_id = cabinet_match.group(1) if cabinet_match else ""
    
    # 提取坐标
    coord_match = re.search(r'坐标\s*[（(](\d+)[,，]\s*(\d+)[）)]', cell_value)
    if coord_match:
        x, y = int(coord_match.group(1)), int(coord_match.group(2))
    else:
        x, y = 0, 0
    
    return cabinet_id, (x, y)

def get_cell_value(worksheet, row: int, col: int) -> Optional[str]:
    """
    获取单元格的值，如果是合并单元格，返回合并单元格的值
    """
    cell = worksheet.cell(row=row, column=col)
    
    # 检查是否在合并单元格范围内
    for merged_range in worksheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            # 返回合并单元格左上角的值
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    
    # 不是合并单元格，返回当前单元格的值
    return cell.value

def find_u_position_range(worksheet, device_row: int, device_col: int, u_col: int) -> Tuple[int, int]:
    """
    查找设备占用的U位范围
    """
    # 检查设备是否在合并单元格中
    for merged_range in worksheet.merged_cells.ranges:
        if (merged_range.min_row <= device_row <= merged_range.max_row and 
            merged_range.min_col <= device_col <= merged_range.max_col):
            # 获取合并单元格的U位范围
            start_u = None
            end_u = None
            
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                u_value = worksheet.cell(row=row, column=u_col).value
                if u_value and str(u_value).isdigit():
                    u_num = int(u_value)
                    if start_u is None or u_num < start_u:
                        start_u = u_num
                    if end_u is None or u_num > end_u:
                        end_u = u_num
            
            return (start_u or 0, end_u or 0)
    
    # 不是合并单元格，返回单行的U位
    u_value = worksheet.cell(row=device_row, column=u_col).value
    if u_value and str(u_value).isdigit():
        return (int(u_value), int(u_value))
    
    return (0, 0)

def process_cabinet(worksheet, cabinet_row: int, cabinet_col: int, room_name: str) -> List[Dict]:
    """
    处理单个机柜，提取其中的所有设备信息
    """
    devices = []
    
    # 获取机柜信息
    cabinet_cell_value = worksheet.cell(row=cabinet_row, column=cabinet_col).value
    cabinet_id, (x, y) = parse_cabinet_info(cabinet_cell_value)
    
    if not cabinet_id:
        return devices
    
    print(f"  处理机柜 {cabinet_id} at ({x},{y})")
    
    # 查找U位列（通常在机柜单元格的右侧）
    u_col = None
    for check_col in range(cabinet_col, min(cabinet_col + 3, worksheet.max_column + 1)):
        # 检查下一行是否有数字（U位）
        check_cell = worksheet.cell(row=cabinet_row + 1, column=check_col)
        if check_cell.value and str(check_cell.value).isdigit():
            u_col = check_col
            break
    
    if not u_col:
        print(f"    警告: 未找到机柜 {cabinet_id} 的U位列")
        return devices
    
    # 从机柜下一行开始扫描
    current_row = cabinet_row + 1
    processed_devices = set()  # 记录已处理的设备，避免重复
    
    while current_row <= worksheet.max_row:
        # 获取U位值
        u_value = worksheet.cell(row=current_row, column=u_col).value
        
        # 如果U位不是数字，说明可能到了下一个机柜
        if not u_value or not str(u_value).isdigit():
            break
        
        u_num = int(u_value)
        
        # 检查左侧是否有设备信息（通常在U位列的左侧）
        device_col = u_col - 1
        device_value = get_cell_value(worksheet, current_row, device_col)
        
        if device_value and str(device_value).strip():
            device_name = str(device_value).strip()
            
            # 避免重复记录同一个设备
            if device_name not in processed_devices:
                processed_devices.add(device_name)
                
                # 获取设备占用的U位范围
                start_u, end_u = find_u_position_range(worksheet, current_row, device_col, u_col)
                
                if start_u == 0:  # 如果没有找到U位范围，使用当前U位
                    start_u = end_u = u_num
                
                devices.append({
                    'new hostname': device_name,
                    'locations': cabinet_id,
                    'U位': f"{start_u}-{end_u}" if start_u != end_u else str(start_u),
                    '机房': room_name,
                    '设备坐标X': x,
                    '设备坐标Y': y
                })
                print(f"    设备: {device_name} (U{start_u}-U{end_u})")
        
        # 如果U位是1，结束当前机柜的扫描
        if u_num == 1:
            break
        
        current_row += 1
    
    # 如果机柜内没有设备，记录空机柜
    if not devices:
        devices.append({
            'new hostname': '空机柜',
            'locations': cabinet_id,
            'U位': '1-52',  # 假设标准机柜是52U
            '机房': room_name,
            '设备坐标X': x,
            '设备坐标Y': y
        })
        print(f"    空机柜")
    
    return devices

def find_all_cabinets(worksheet, room_name: str) -> List[Dict]:
    """
    在工作表中查找所有机柜并提取设备信息
    """
    all_devices = []
    processed_cabinets = set()  # 记录已处理的机柜位置，避免重复
    
    # 遍历所有单元格，查找包含"机柜"的单元格
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            
            # 查找包含"机柜"关键字的单元格
            if cell_value and isinstance(cell_value, str) and "机柜" in cell_value:
                # 避免处理重复的机柜
                cabinet_pos = (row, col)
                if cabinet_pos not in processed_cabinets:
                    processed_cabinets.add(cabinet_pos)
                    
                    # 处理这个机柜
                    devices = process_cabinet(worksheet, row, col, room_name)
                    all_devices.extend(devices)
    
    return all_devices

def read_cabinet_info(excel_file: str, output_file: str = None):
    """
    读取Excel文件中的机柜信息并导出
    """
    # 加载工作簿
    print(f"正在加载文件: {excel_file}")
    wb = load_workbook(excel_file, data_only=True)
    
    all_devices = []
    
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        print(f"\n处理工作表: {sheet_name}")
        worksheet = wb[sheet_name]
        
        # 查找并处理所有机柜
        devices = find_all_cabinets(worksheet, sheet_name)
        all_devices.extend(devices)
        
        print(f"  在 {sheet_name} 中找到 {len(devices)} 个设备/机柜")
    
    # 关闭工作簿
    wb.close()
    
    # 创建DataFrame
    df = pd.DataFrame(all_devices)
    
    # 按机房、机柜位置排序
    if not df.empty:
        # 提取机柜字母和数字部分用于排序
        df['cabinet_letter'] = df['locations'].str.extract(r'([A-Z]+)')
        df['cabinet_number'] = df['locations'].str.extract(r'(\d+)').astype(int)
        df = df.sort_values(['机房', 'cabinet_letter', 'cabinet_number'])
        df = df.drop(['cabinet_letter', 'cabinet_number'], axis=1)
    
    # 输出文件名
    if not output_file:
        base_name = os.path.splitext(excel_file)[0]
        timestamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
        output_file = f"{base_name}_设备信息_{timestamp}.xlsx"
    
    # 保存到Excel文件
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='设备信息')
        
        # 调整列宽
        worksheet = writer.sheets['设备信息']
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    print(f"\n处理完成!")
    print(f"总共找到 {len(all_devices)} 条设备/机柜信息")
    print(f"结果已保存到: {output_file}")
    
    # 显示统计信息
    if not df.empty:
        print("\n统计信息:")
        print(f"- 机房数量: {df['机房'].nunique()}")
        print(f"- 机柜数量: {df['locations'].nunique()}")
        print(f"- 设备数量: {len(df[df['new hostname'] != '空机柜'])}")
        print(f"- 空机柜数量: {len(df[df['new hostname'] == '空机柜'])}")
        
        print("\n各机房设备分布:")
        for room in df['机房'].unique():
            room_df = df[df['机房'] == room]
            device_count = len(room_df[room_df['new hostname'] != '空机柜'])
            cabinet_count = room_df['locations'].nunique()
            print(f"  {room}: {cabinet_count} 个机柜, {device_count} 个设备")
    
    return df

def main():
    """
    主函数
    """
    # 输入文件路径
    input_file = "机柜图.xlsx"  # 请根据实际情况修改文件路径
    
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到文件 {input_file}")
        print("请确保文件在当前目录下，或提供完整路径")
        return
    
    try:
        # 读取机柜信息
        df = read_cabinet_info(input_file)
        
        # 显示部分结果预览
        if not df.empty:
            print("\n结果预览 (前10行):")
            print(df.head(10).to_string())
        
    except Exception as e:
        print(f"\n发生错误: {e}")
        print("\n详细错误信息:")
        import traceback
        traceback.print_exc()
        
        print("\n请检查:")
        print("1. Excel文件格式是否正确")
        print("2. 机柜信息格式是否为: '25KW 机柜(A01)坐标(1,2)'")
        print("3. U位信息是否在机柜名称的下方")
        print("4. 设备信息是否在U位的左侧")

if __name__ == "__main__":
    main()

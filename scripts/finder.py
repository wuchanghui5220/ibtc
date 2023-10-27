import pandas as pd
import openpyxl
import sys
import os
import time
import random
import re
import glob
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from termcolor import colored  


files = {
    'HW': [],
    'SWS X': [],
    'SWS P': [] 
}
hw_files = glob.glob('*HW*.xlsx')  
for f in hw_files:
    files['HW'].append(f)
sws_x_files = glob.glob('*SWS X*.xlsx')  
for f in sws_x_files:
    files['SWS X'].append(f)   
sws_p_files = glob.glob('*SWS P*.xlsx')
for f in sws_p_files:
    files['SWS P'].append(f)
# print(files)
file1 = files['HW'][0]
file2 = files['SWS X'][0]
file3 = files['SWS P'][0]

template_file = 'template.xlsx'
if not os.path.exists(file1):
    print("Excel file does not exist")
    sys.exit(1)
def search_sws_x(excel_file, temp_search_pn):
    last_color = None
    colors = ['white', 'green','yellow','blue','magenta','cyan']
    keyword = temp_search_pn
    df = pd.read_excel(excel_file, skiprows=3)
    index = df[df['Configured SKU: Product Name']==keyword].index[0]
    result = df.loc[index:len(df)] 
    for i in range(index+1, len(df)):
        if pd.notna(df.iloc[i,0]):
            result = result.loc[index:i-1]
            break
    result = result.reset_index(drop=True)
    print(result)
    while True:
        selected_row = input("(直接回车则不选择)请选择对应服务Code的编号: ")
        if selected_row == "":
            return None
        elif selected_row.isdigit():
            selected_row = int(selected_row)
            if 0 <= selected_row < len(result):
                selected_code = result.iloc[selected_row, 3]
                print(selected_code)
                return selected_code
                break
            else:
                print("输入有误，请重新输入！")
        else:
            print("输入有误，请重新输入！")
def search_sws(excel_file,x_pn,qty):
    last_color = None
    colors = ['white', 'green','yellow','blue','magenta','cyan']
    x_pn = x_pn.upper()
    # print(f'x_pn: {x_pn}')
    # print("type: ")
    # print(type(x_pn))
    need_PN = []
    dfsws = pd.read_excel(excel_file) 
    search_list = ['Price Book Organizer', 'Material', 'Mellanox Legacy', 'Description','P8 - NPN PRICE']
    dfsws = dfsws.loc[:,search_list]
    dfsws_pn = dfsws
    # search_result = dfsws[dfsws['Material'].str.contains(x_pn, na=False)]

    search_result = dfsws[dfsws['Mellanox Legacy'].str.contains(x_pn, na=False)]
    if search_result.empty:
        search_result = dfsws[dfsws['Material'].str.contains(x_pn, na=False)]

    if search_result.empty:
        print("未找到结果!sws")
    else:
        max_len = len(max(search_result['Mellanox Legacy'], key=len))
        for i, row in enumerate(search_result[['Mellanox Legacy','Description', 'P8 - NPN PRICE']].iterrows()):
            while True:
                c = random.choice(colors)
                if c != last_color:
                    break
            last_color = c
            if i <= 9:
                name = row[1]['Mellanox Legacy'].ljust(max_len + 1)  
                desc = row[1]['Description']
                price = row[1]['P8 - NPN PRICE']
                # print(f"{i}. {name} {desc}  ${price}")
                print(colored(f"{i}. {name} {desc} ${price} ", c))
            elif i > 9:
                name = row[1]['Mellanox Legacy'].ljust(max_len)  
                desc = row[1]['Description']
                price = row[1]['P8 - NPN PRICE']
                # print(f"{i}. {name} {desc}  ${price}")
                print(colored(f"{i}. {name} {desc} ${price} ", c))
        while True:
            try:
                choice = input("请选择对应服务的序号(直接回车默认选择 '0' ):")
                if choice == '':
                    selectedsws = search_result.iloc[0]['Mellanox Legacy']
                    need_PN.append(selectedsws)
                    # 提示输入数量
                    # qtysws = int(input("请输入{}的购买数量:".format(selectedsws)))
                    dfsws_pn = dfsws_pn[dfsws_pn['Mellanox Legacy'].isin(need_PN)]
                    df_new = dfsws_pn.copy()
                    df_new.loc[:, 'Brand'] = 'NVIDIA' 
                    df_new.loc[:, 'Quality'] = qty
                    break
                
                choice = int(choice)
                if choice >= 0 and choice < len(search_result):
                    selectedsws = search_result.iloc[choice]['Mellanox Legacy']
                    need_PN.append(selectedsws)
                    # 提示输入数量
                    # qtysws = int(input("请输入{}的购买数量:".format(selectedsws)))
                    dfsws_pn = dfsws_pn[dfsws_pn['Mellanox Legacy'].isin(need_PN)]
                    df_new = dfsws_pn.copy()
                    df_new.loc[:, 'Brand'] = 'NVIDIA' 
                    df_new.loc[:, 'Quality'] = qty
                    break
                else:
                    print("没有这个选项,请重新输入")
                    
            except ValueError:
                print("请输入一个数字!")
            except KeyboardInterrupt:
                break
        df_new = df_new.drop('Material', axis=1)
        # print(df_new)
        if not df_new.empty:
            return df_new
def search_and_select(excel_file):
    dfh = pd.read_excel(excel_file) 
    dfh = dfh.loc[:,['Price Book Organizer', 'Product Name','Product Description','P3 = OEM Price']]
    last_color = None
    colors = ['white', 'green','yellow','blue','magenta','cyan']
    # search_columns = ['Product Code', 'Product Name', 'Product Description']
    # search_columns = dfh.columns
    while True:
        search_term = input("请输入搜索词:")
        search_term = search_term.upper()
        # search_result = dfh[dfh['Product Name'].str.contains(search_term)]
        search_result = dfh[dfh['Product Name'].str.contains(search_term)]
        if search_result.empty:
            print("未找到结果,请重新输入!")
        else:
            break
    
    max_len = len(max(search_result['Product Name'], key=len))
    
    for i, row in enumerate(search_result[['Product Name','Product Description', 'P3 = OEM Price']].iterrows()):
        while True:
            c = random.choice(colors)
            if c != last_color:
                break
        last_color = c
        if i <= 9:
            name = row[1]['Product Name'].ljust(max_len + 1)  
            desc = row[1]['Product Description']
            price = row[1]['P3 = OEM Price']
            print(colored(f"{i}. {name} {desc} ${price} ", c))
        elif i > 9:
            name = row[1]['Product Name'].ljust(max_len)  
            desc = row[1]['Product Description']
            price = row[1]['P3 = OEM Price']
            # print(f"{i}. {name} {desc} ${price} [{i}]")
            print(colored(f"{i}. {name} {desc} ${price} ", c))

    need_PN = []   
    while True:
        try:
            choice = input("请输入要选择的序号(直接回车结束本次选择):")
            if choice == '':
                break
            
            choice = int(choice)
            if choice >= 0 and choice < len(search_result):
                selected = search_result.iloc[choice]['Product Name']
                need_PN.append(selected)
                # selected_PN.append(selected)
                # 提示输入数量
                qty = int(input("请输入{}的购买数量:".format(selected)))
                # 数量_dict[selected] = 数量 
                # if selected in qty_dict:
                #     qty_dict[selected] += qty
                # else:
                #     qty_dict[selected] = qty
                df_pn = dfh

                df_pn = df_pn[df_pn['Product Name'].isin(need_PN)]
                df_new = df_pn.copy()
                # 在返回的DataFrame最后加一列Quality
                df_new.loc[:, 'Brand'] = 'NVIDIA' 
                df_new.loc[:, 'Quality'] = qty
                break
            else:
                print("没有这个选项,请重新输入")
                
        except ValueError:
            print("请输入一个数字!")
        except KeyboardInterrupt:
            break
    
    print("本次选择的产品和数量:")      
    print(f'{need_PN[0]} : {qty}')
    # print()
    # print(df_new)
    return df_new, qty
def search_and_select_sws(excel_file):
    dfsws = pd.read_excel(excel_file) 
    dfsws = dfsws.loc[:,['Price Book Organizer', 'Mellanox Legacy', 'Description','P8 - NPN PRICE']]
    while True:
        search_term = input("请输入搜索词:")
        search_term = search_term.upper()
        search_result = dfsws[dfsws['Mellanox Legacy'].str.contains(search_term, na=False)]

        if search_result.empty:
            print("未找到结果,请重新输入!")
        else:
            break
    
    max_len = len(max(search_result['Mellanox Legacy'], key=len))
    
    for i, row in enumerate(search_result[['Mellanox Legacy','Description']].iterrows()):
        if i <= 9:
            name = row[1]['Mellanox Legacy'].ljust(max_len + 1)  
            desc = row[1]['Description']
            print(f"{i}. {name} {desc}  [{i}]")
        elif i > 9:
            name = row[1]['Mellanox Legacy'].ljust(max_len)  
            desc = row[1]['Description']
            print(f"{i}. {name} {desc}  [{i}]")

    need_PN = []   
    while True:
        try:
            choice = input("请输入要选择的序号(直接回车结束本次选择):")
            if choice == '':
                break
            
            choice = int(choice)
            if choice >= 0 and choice < len(search_result):
                selectedsws = search_result.iloc[choice]['Mellanox Legacy']
                need_PN.append(selectedsws)
                # 提示输入数量
                qtysws = int(input("请输入{}的购买数量:".format(selectedsws)))
                # 数量_dict[selected] = 数量 
                # if selectedsws in lic_qty_dict:
                #     lic_qty_dict[selectedsws] += qtysws
                # else:
                #     lic_qty_dict[selectedsws] = qtysws
                # dfsws_pn = dfsws

                dfsws = dfsws[dfsws['Mellanox Legacy'].isin(need_PN)]
                df_new = dfsws.copy()
                # 在返回的DataFrame最后加一列Quality
                df_new.loc[:, 'Brand'] = 'NVIDIA' 
                df_new.loc[:, 'Quality'] = qtysws
                # df_new['Brand'] = 'NVIDIA'
                # df_new['Quality'] = qtysws
            else:
                print("没有这个选项,请重新输入")         
        except ValueError:
            print("请输入一个数字!")
        except KeyboardInterrupt:
            break
    # need_PN = list(lic_qty_dict.keys())
    print("已选择的license和数量:")      
    print(need_PN, qtysws)
    print("打印dfsws")
    print(df_new)
    return df_new, qtysws
def get_ex_rate(default=7.35):
    rate = input(f'输入人民币对美元汇率(默认值{default}): ') or default
    try:
        return float(rate)
    except ValueError:
        print("输入无效, 使用默认值")
        return default
def get_npn_rate(prompt, default=10):
    # rate = int(input(prompt) or default)
    rate = float(rate / 100) or default
    try:
        return rate
    except ValueError:
        print("输入无效, 使用默认值")
        return default

# start

current_time = time.strftime("%Y%m%d-%H%M%S")
first_file = f'df-{current_time}.xlsx'

################################################################ 主程序开始
while True:
    df_data, qty = search_and_select(file1)

    # 首次写入,创建新工作簿
    if not os.path.exists(first_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Price Book Organizer', 'Product Name','Product Description','P3 = OEM Price', 'Brand' ,'Quality'])
    # 后续写入，加载工作簿
    else:
        wb = openpyxl.load_workbook(first_file)
        ws = wb.active
    # 获取最后一行号
    last_row = ws.max_row 
    # for r in excel.dataframe_to_rows(df_data, index=False, header=False):
    for index, row in df_data.iterrows():
        ws.append(row.tolist())
    # 保存工作簿 
    if wb:
        wb.save(first_file)

    # 获取最后一行号
    last_row = ws.max_row
    # 读取服务查询file2
    temp_search_pn = df_data.iloc[0, 1] 
    # print("打印已选择的PN")
    # print(temp_search_pn)

    x_pn = search_sws_x(file2, temp_search_pn)
    # print("打印搜索到的x_pn")
    # print(x_pn)
    if x_pn != None:
        sws_price_df = search_sws(file3, x_pn,qty)
        # print("显示是否查询到sws的df数据：")
        # print(sws_price_df)
        if not sws_price_df.empty:
            wb = openpyxl.load_workbook(first_file)
            ws = wb.active
            # 获取最后一行号
            last_row = ws.max_row 
            # for r in excel.dataframe_to_rows(df_data, index=False, header=False):
            for index, row in sws_price_df.iterrows():
                ws.append(row.tolist())
            # 保存工作簿 
            if wb:
                wb.save(first_file)
            print("qty:"+ str(qty))
    again = input("是否需要再次搜索(默认y/Y)?")
    if again.upper() == 'N':
        break

# license  order
license_PN = []
buy_lic = input('''是否购买UFM或者RiverMAX软件license？
    [不够买license，默认n/N] 
    [输入 l/L 查询license]
请输入你的选择： ''')

if buy_lic.upper() == 'L':
    dfsws, qtysws = search_and_select_sws(file3)
    if not dfsws.empty:
        wb = openpyxl.load_workbook(first_file)
        ws = wb.active
        # 获取最后一行号
        last_row = ws.max_row 
        # for r in excel.dataframe_to_rows(df_data, index=False, header=False):
        for index, row in dfsws.iterrows():
            ws.append(row.tolist())
        # 保存工作簿 
        if wb:
            wb.save(first_file)
            print("qtysws:"+ str(qtysws))
elif buy_lic == '':
    print("用户选择不需要购买license！\n")
    pass
else:
    print("用户选择不需要购买license！\n")
    pass

df = pd.read_excel(first_file)

# 在'P3 = OEM Price'后插入新列'汇率'
ER = get_ex_rate()
df.insert(loc=6, column='汇率', value=ER) 

# 在'汇率'后插入新列'税费'
df.insert(loc=7, column='税费', value=0.13)

# 在'税费'后插入新列'总代利润点' 
df.insert(loc=8, column='总代利润点', value=0.05)

# 在'总代利润点'后插入新列'NPN利润点'
df.insert(loc=9, column='NPN利润点', value=None)

all_nr = input('请输入统一的利润点（直接回车将逐个输入）： ')
if all_nr == '':
    for i, row in df.iterrows():
        product = row['Product Name']
        NR = get_npn_rate("请输入产品{}的利润点(默认10个点):".format(product), 10)
        df.at[i, 'NPN利润点'] = NR
else:

    for i, row in df.iterrows():
        product = row['Product Name']
        NR = float(all_nr) / 100
        df.at[i, 'NPN利润点'] = NR


# 在'数量'后插入新列'含税单价', 计算含税单价
df.insert(loc=10, column='含税单价', value=(df['P3 = OEM Price'] * df['汇率'] * (1 + df['税费']) * (1 + df['总代利润点']) * (1 + df['NPN利润点'])).astype(int))

# 在'含税单价'后插入新列'含税总价', 计算含税总价
df.insert(loc=11, column='含税总价', value=df['Quality'] * df['含税单价'])

xlsx_file_name = f'正阳恒卓-{current_time}.xlsx'
df.to_excel(xlsx_file_name, index=False)

wb = openpyxl.Workbook()
ws = wb.active

for index, row in df.iterrows():
    ws.append(row.tolist())

# 设置D列为美元格式 
for i in range(1, ws.max_row+1):
  c_col = 'D' + str(i)
  ws[c_col].number_format = u'$#,##0.00' 



# 设置 F，G, H 列为百分比 % 格式
for i in range(1, ws.max_row+1):
  e_col = 'H' + str(i)
  f_col = 'I' + str(i) 
  g_col = 'J' + str(i)
  
  ws[e_col].number_format = '0%'
  ws[f_col].number_format = '0%'
  ws[g_col].number_format = '0%'

# 设置 'Quality' 这一列为整数
for i in range(1, ws.max_row+1):
  c_col = 'F' + str(i)
  ws[c_col].number_format = '0' 

# 设置J，K 列为货币RMB格式
for i in range(1, ws.max_row+1):
    i_col = 'K' + str(i) 
    j_col = 'L' + str(i)
    ws[i_col].number_format = u'¥#,##0.00'
    ws[j_col].number_format = u'¥#,##0.00' 

# 在第一列前插入新列,创建序号
ws.insert_cols(1)

# 插入序号
for i in range(1, ws.max_row+1):
    c = ws.cell(i, 1) 
    c.value = i
    c.number_format = '0' # 指定数字格式为General "常规"

# 获取最后一列的索引 
last_col = ws.max_column 

# 在最后一列后插入新列
ws.insert_cols(last_col+1) 

# 在新插入的这一列写入值
Delivery_date = input("请输入货期(ex:2-3)周：")
if Delivery_date == "":
    Delivery_date = "2-3"

for cell in ws["N"]:
    cell.value = Delivery_date + "周"

# 获取表格的所有单元格
last_col = get_column_letter(ws.max_column) 
cells = ws[f'A1:{last_col}{ws.max_row}']
# 表格的所有单元格垂直居中，水平居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center')

# 要实现左对齐的列号, 开启自动换行
columns = ['B', 'D']

for column in columns:
    for cell in ws[column]:
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

# 在第一列前插入新列，创建一列空列，以便后面从第2列复制带有边框的数据表
ws.insert_cols(1)
# 添加边框
thin = Side(border_style="thin", color="000000")
for row in ws.rows:
    for cell in row:
        if cell.value:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
# 设置列宽
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 21
ws.column_dimensions['D'].width = 21
ws.column_dimensions['E'].width = 81
ws.column_dimensions['F'].width = 21
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 18
ws.column_dimensions['N'].width = 18
ws.column_dimensions['O'].width = 10

xlsx_file_name_new = f'正阳恒卓报价单-留存单-{current_time}.xlsx'
wb.save(xlsx_file_name_new)

# 打开源文件和目标文件
source_workbook = openpyxl.load_workbook(xlsx_file_name_new)
target_workbook = openpyxl.load_workbook(template_file)

# 选择源文件和目标文件中的工作表
source_sheet = source_workbook.active
target_sheet = target_workbook.active

# 选择要复制的源数据范围
source_columns = ['B', 'C', 'D', 'E', 'G', 'H', 'M', 'N', 'O']
source_max_row = source_sheet.max_row

# 选择目标文件的要粘贴的位置（B16开始）
target_start_row = 16
target_start_column = 'B'


# 复制源文件的数据到目标文件
for source_column in source_columns:
    for row_index in range(1, source_max_row + 1):
        source_cell = source_sheet[source_column + str(row_index)].value
        target_row = row_index + target_start_row - 1
        target_column = chr(ord(target_start_column) + source_columns.index(source_column))
        target_sheet[target_column + str(target_row)] = source_cell



# 另存为新文件
new_filename = f'正阳恒卓报价单-{current_time}.xlsx'
target_workbook.save(new_filename)



# 关闭工作簿
source_workbook.close()
target_workbook.close()

# 打开new_filename文件
workbook = openpyxl.load_workbook(new_filename)

# 选择要操作的工作表
sheet = workbook.active

# 修改J10单元格的内容
current_date = time.strftime("%Y-%m-%d")
sheet['J10'] = f"报价日期：{current_date}"  # 将"新的内容"替换为你想要设置的新内容

# 保存修改后的文件
workbook.save(new_filename)

# 关闭工作簿
workbook.close()

print(f"报价完成，数据已保存到 [{new_filename}] 文件中, 请及时查看和检查！\n")

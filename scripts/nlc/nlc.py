import openpyxl


# 标准U高度
U_HEIGHT = 4.45  # 单位是cm
# 机柜的总U高度
TOTAL_U_HEIGHT = 42
# 机柜底部到地板下方线槽的距离+ 拐弯+美观长度预留
EXTRA_CONDUIT_HEIGHT = 180  # 单位是cm
# 排间距离（从一排的后面到另一排的前面）
ROW_DISTANCE = 180  # 单位是cm
# 机柜深度的一半
HALF_CABINET_DEPTH = 60  # 单位是cm
# 机柜的宽度为80cm
CABINET_WIDTH = 160
# 纵向线槽所在的列，可以根据不同机房调整
conduit_columns = [3, 11]  # 例如，A列，K列

def parse_data_to_dicts(file_name):
    devices = {}
    cabinet_locations = {}
    with open(file_name, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()  # 去掉换行符
            cabinet, row, column = line.split('-')  # 分割字符串
            # 处理机柜位置字典
            if cabinet not in cabinet_locations:
                cabinet_row = ord(cabinet[0]) - ord('A') + 1  # 使用新变量来避免覆盖原始的row
                cabinet_col = int(cabinet[1:])
                cabinet_locations[cabinet] = (cabinet_row, cabinet_col)
            # 处理设备字典
            device_name = f"{cabinet}-{row}-{column}"
            devices[device_name] = (cabinet, int(row), int(column))
    return devices, cabinet_locations

# 调用函数并获取字典结果
devices, cabinet_locations = parse_data_to_dicts('c_column_data7.txt')
# 打印字典结果
print("设备字典:")
for device, location in devices.items():
    print(f"{device}: {location}")
print("\n机柜位置字典:")
for cabinet, location in cabinet_locations.items():
    print(f"{cabinet}: {location}")


# 计算设备到机柜地板下方线槽的垂直距离
def calculate_vertical_distance(u_position):
    return u_position * U_HEIGHT + EXTRA_CONDUIT_HEIGHT

# 计算不同排机柜间设备的线缆长度
def calculate_different_rows_distance(device1, device2):
    # 获取设备的机柜位置和U位置
    cabinet1, u_position1, device_u_height1 = devices[device1]
    cabinet2, u_position2, device_u_height2 = devices[device2]
    # 获取机柜的排和列位置
    row1, col1 = cabinet_locations[cabinet1]
    row2, col2 = cabinet_locations[cabinet2]
    # 计算设备底部到机柜的垂直距离
    vertical_distance1 = calculate_vertical_distance(u_position1)
    vertical_distance2 = calculate_vertical_distance(u_position2)
    # 计算两排之间的垂直距离（包括机柜深度）
    vertical_distance_between_rows = abs(row1 - row2) * (ROW_DISTANCE + HALF_CABINET_DEPTH * 2)
    # print(f'排间距离：{vertical_distance_between_rows}')
    # 处于中间的情况
    if col1 in range(conduit_columns[0], conduit_columns[1] + 1) and col2 in range(conduit_columns[0], conduit_columns[1] + 1):
        # 计算两个纵向线槽之间的水平距离（逆时针或顺时针路径中的较短一条）
        horizontal_distance_case1 = abs((col1 + col2) - 2 * conduit_columns[0]) * CABINET_WIDTH # 靠近 conduit_columns[0]
        horizontal_distance_case2 = abs(2 * conduit_columns[1] - (col1 + col2)) * CABINET_WIDTH # 靠近 conduit_columns[1]
        # 计算总线缆长度，逆时针方向
        case1_cable_length = (
            vertical_distance1 +
            horizontal_distance_case1 +
            vertical_distance_between_rows +
            vertical_distance2)

        # 计算总线缆长度，顺时针方向
        case2_cable_length = (
            vertical_distance1 +
            horizontal_distance_case2 +
            vertical_distance_between_rows +
            vertical_distance2)
 
        # 返回较短的线缆长度
        min_length = min(case1_cable_length, case2_cable_length)
        if case1_cable_length == min_length:
            print(f'''{device1} <--> {device2} : {int(min_length)}cm
            处于2个纵向线槽之间=|=.=.=|=，逆时针方向
            本端垂直线长: {vertical_distance1}cm 
            两端水平线长: {horizontal_distance_case1}cm 
            两端垂直线长: {vertical_distance_between_rows}cm 
            对端垂直线长: {vertical_distance2}cm''')
        elif case2_cable_length == min_length:
            print(f'''{device1} <--> {device2} : {int(min_length)}cm
            处于2个纵向线槽之间=|=.=.=|=，顺时针方向
            本端垂直线长: {vertical_distance1}cm 
            两端水平线长: {horizontal_distance_case2}cm 
            两端垂直线长: {vertical_distance_between_rows}cm 
            对端垂直线长: {vertical_distance2}cm''')
        return str(int(min_length))

    # 都小于conduit_columns[0]
    elif col1 in range(1, conduit_columns[0] + 1) and col2 in range(1, conduit_columns[0] + 1):
        horizontal_distance= abs(2 * conduit_columns[0] - (col1 + col2)) * CABINET_WIDTH # 全部居左侧于 conduit_columns[0]
        # print(f'第3列线槽左侧水平距离 case:{horizontal_distance}')
        # 计算总线缆长度（conduit_columns[0]）
        case_cable_length = (
            vertical_distance1 +
            horizontal_distance +
            vertical_distance_between_rows +
            vertical_distance2)

        print(f'''{device1} <--> {device2} : {str(int(case_cable_length))}cm
            两端处于1个纵向线槽同侧=.=.=|=：
            本端垂直线长: {vertical_distance1}cm 
            两端水平线长: {horizontal_distance}cm 
            两端垂直线长: {vertical_distance_between_rows}cm 
            对端垂直线长: {vertical_distance2}cm''')

        # 返回线缆长度
        return str(int(case_cable_length))

    # 处于conduit_columns[0]两侧情况1
    elif col1 in range(1, conduit_columns[0] + 1) and col2 in range(conduit_columns[0], conduit_columns[1] + 1):
        # 计算两个纵向线槽之间的水平距离
        horizontal_distance_case2 = abs(col2 - col1) * CABINET_WIDTH # conduit_columns[0]两侧

        # 计算总线缆长度
        case2_cable_length = (
            vertical_distance1 +
            horizontal_distance_case2 +
            vertical_distance_between_rows +
            vertical_distance2)

        print(f'''{device1} <--> {device2} : {str(int(case2_cable_length))}cm
            两端处于1个纵向线槽两侧=.=|=.=：
            本端垂直线长: {vertical_distance1}cm 
            两端水平线长: {horizontal_distance_case2}cm 
            两端垂直线长: {vertical_distance_between_rows}cm 
            对端垂直线长: {vertical_distance2}cm''')

        # 返回较短的线缆长度
        return str(int(case2_cable_length))
    # 处于conduit_columns[0]两侧情况2
    elif col2 in range(1, conduit_columns[0] + 1) and col1 in range(conduit_columns[0], conduit_columns[1] + 1):
        # 计算两个纵向线槽之间的水平距离
        horizontal_distance_case2_2 = abs(col2 - col1) * CABINET_WIDTH # conduit_columns[0]两侧

        # 计算总线缆长度
        case2_2_cable_length = (
            vertical_distance1 +
            horizontal_distance_case2_2 +
            vertical_distance_between_rows +
            vertical_distance2
        )
        print(f'''{device1} <--> {device2} : {str(int(case2_2_cable_length))}cm
            两端处于1个纵向线槽两侧=.=|=.=：
            本端垂直线长: {vertical_distance1}cm 
            两端水平线长: {horizontal_distance_case2_2}cm 
            两端垂直线长: {vertical_distance_between_rows}cm 
            对端垂直线长: {vertical_distance2}cm''')

        # 返回较短的线缆长度
        return str(int(case2_2_cable_length))

def determine_cable_length_category(length_cm):
    length_cm = int(length_cm)
    # 定义区间值（以厘米为单位）
    intervals = [500, 1000, 1500, 2000, 3000, 5000, 10000]
    labels = ['5m', '10m', '15m', '20m', '30m', '50m', '100m']
    
    for i, interval in enumerate(intervals):
        if length_cm <= interval:
            return labels[i]
    
    # 如果长度超过了最大区间，返回最大的标签
    return labels[-1]
    
def calculate_cable_length(formatted_cell_d, device1, device2):
    # 检查设备键是否存在于字典中
    # print("Device 1:", device1)
    # print("Device 2:", device2)
    
    # 确保device1和device2在devices字典中
    if device1 not in devices or device2 not in devices:
        print(f"Error: Device '{device1}' or '{device2}' not found in devices dictionary.")
        return 0  # 返回一个默认的长度值或者抛出一个异常
    
    cabinet1, u_position1, device_u_height1 = devices[device1]
    cabinet2, u_position2, device_u_height2 = devices[device2]

    # 计算设备底部到机柜的垂直距离
    vertical_distance1 = calculate_vertical_distance(u_position1)
    vertical_distance2 = calculate_vertical_distance(u_position2)
        
    if cabinet1 == cabinet2:
        # 同一机柜内线缆长度计算，或者直接等于固定值，比如3米
        same_cabinet_length = abs(u_position1 - u_position2) * U_HEIGHT + EXTRA_CONDUIT_HEIGHT
        length = str(int(same_cabinet_length))
    elif cabinet_locations[cabinet1][0] == cabinet_locations[cabinet2][0]:
        # 同一排不同机柜间线缆长度计算
        horizontal_distance = abs(cabinet_locations[cabinet1][1] - cabinet_locations[cabinet2][1]) * CABINET_WIDTH
        # horizontal_distance = (calculate_same_row_distance(cabinet1, cabinet2) +
        #           (u_position1 + u_position2) * U_HEIGHT + EXTRA_CONDUIT_HEIGHT)
        print(f'''{device1} <--> {device2} : {str(int(vertical_distance1 + horizontal_distance + vertical_distance2))}cm
            两端处于同一排不同机柜=.=.=：
            本端垂直线长: {vertical_distance1}cm 
            两端水平线长: {horizontal_distance}cm 
            两端垂直线长: 0cm 
            对端垂直线长: {vertical_distance2}cm''')
        length = str(int(vertical_distance1 + horizontal_distance + vertical_distance2))
    else:
        # 不同排机柜间线缆长度计算
        distance = calculate_different_rows_distance(device1, device2)
        if distance is None:
            print(f"Error: Distance calculation returned None for devices '{device1}' and '{device2}'.")
            return 0  # 返回一个默认的长度值或者抛出一个异常
        length = distance

    # 打印结果供调试# 返回线缆长度，以默认 cm 单位
    # return f'{formatted_cell_d} : {length}cm'}

    # 返回线缆长度，以 m 为单位
    # 计算区间值
    length_category = determine_cable_length_category(length)
    return f'{formatted_cell_d} : {length}cm : {length_category}'




# 示例：计算两台设备间的线缆长度
# cable_length = calculate_cable_length("F05-6U-02", "F10-6U-04")
# print(f"The cable length between the two devices is: {cable_length} cm")


# 加载Excel文件
workbook = openpyxl.load_workbook('demo.xlsx')

# 获取Sheet1工作表
worksheet = workbook['Sheet3']
# 迭代行数据,从第3594行开始
for row in worksheet.iter_rows(min_row=3953):
    # 获取C列单元格
    cell_c = row[2]
    
    # 检查C列单元格是否为合并单元格
    merged_value = None
    for merged_range in worksheet.merged_cells:
        if cell_c.coordinate in merged_range:
            # 获取合并单元格的值
            merged_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            break
    else:
        merged_value = cell_c.value
    
    # 获取G列的单元格值
    cell_g = row[6]
    # 获取 G 列单元格的实际数据值
    actual_cell_g_value = None
    if isinstance(cell_g.value, str) and cell_g.value.startswith('='):
        # 如果单元格值是以 '=' 开头的字符串,则表示它是一个公式或引用
        # 在这种情况下,我们需要获取被引用单元格的值
        referenced_cell = worksheet[cell_g.value[1:]]  # 获取被引用的单元格对象
        actual_cell_g_value = referenced_cell.value  # 获取被引用单元格的值
    elif cell_g.data_type == 'f':
        # 如果单元格包含公式,则使用单元格的 value 属性获取公式的计算结果
        actual_cell_g_value = cell_g.value
    elif cell_g.data_type == 'n':
        # 如果单元格包含数字,则直接使用单元格的 value 属性
        actual_cell_g_value = cell_g.value
    else:
        # 处理其他类型的数据
        actual_cell_g_value = str(cell_g.value)

    
    # 获取D列的单元格值
    cell_d = row[3].value

    # 如果G列的数据以 ★ 开头,则跳过该行
    # 如果G列的单元格为空，或者G列的数据以 '★' 开头，则跳过该行
    if (not actual_cell_g_value) or (str(actual_cell_g_value).startswith('★')):
        continue

    # if cell_g and str(cell_g).startswith('★'):
    #     continue

    # 如果C列、G列和D列都有数据,则删除开头的 ☆ 符号并拼接字符串
    elif merged_value and cell_g and cell_d:
        # 删除开头的 ☆ 符号
        formatted_merged_value = str(merged_value).lstrip('☆')
        formatted_cell_g = str(actual_cell_g_value).lstrip('☆')
        formatted_cell_d = str(cell_d).lstrip('☆')
        print("")
        # 调用计算两台设备间的线缆长度函数
        result = calculate_cable_length(formatted_cell_d, formatted_merged_value, formatted_cell_g)
        
        # 将结果保存到J列
        row[9].value = result
        
# 保存修改后的Excel文件
workbook.save('demo9_6.xlsx')

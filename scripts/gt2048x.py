import math
from openpyxl import Workbook

def generate_flexible_topology_256(num_servers, ports_per_switch=64):
    # 计算需要的SU数量和Leaf交换机数量
    num_sus = math.ceil(num_servers / 32)
    num_leafs = num_sus * 8
    
    # 计算Spine交换机数量（必须是32的因数）
    spine_factors = [2, 4, 8, 16, 32]
    num_spines = next(factor for factor in spine_factors if factor >= num_leafs / 2)

    # 初始化连接列表
    leaf_spine_connections = []
    server_leaf_connections = []

    # 计算每个Leaf到Spine的连接数
    connections_per_spine = 32 // num_spines

    # 生成Leaf到Spine的连接
    for leaf_index in range(num_leafs):
        su_index = leaf_index // 8
        local_leaf_index = leaf_index % 8
        leaf_name = f"LG{su_index+1}-Leaf{local_leaf_index+1}"
        
        for spine_index in range(num_spines):
            spine_name = f"Spine{spine_index+1}"
            for port_offset in range(connections_per_spine):
                leaf_port = spine_index * connections_per_spine + port_offset + 1
                spine_port = leaf_index * connections_per_spine + port_offset + 1
                leaf_spine_connections.append((leaf_name, f"Port{leaf_port}", spine_name, f"Port{spine_port}"))

    # 生成服务器到Leaf的连接
    for server_index in range(num_servers):
        su_index = server_index // 32
        local_server_index = server_index % 32
        gpu_index = local_server_index + 1

        for hca_index in range(1, 9):  # 每个GPU有8个HCA
            leaf_index = hca_index - 1
            leaf_port = 33 + local_server_index

            server_name = f"SU{su_index+1}-GPU{gpu_index}"
            hca_port = f"HCA{hca_index}"
            leaf_name = f"LG{su_index+1}-Leaf{leaf_index+1}"
            server_leaf_connections.append((server_name, hca_port, leaf_name, f"Port{leaf_port}"))

    return leaf_spine_connections + server_leaf_connections

def generate_three_tier_topology_512(total_servers, num_cgs=8, leaves_per_lg=8, spines_per_sg=8, cores_per_cg=8, ports_per_leaf=64, ports_per_spine=64):
    # 根据服务器总数计算LG和SG的数量
    servers_per_lg = 32  # 假设每个LG有32台服务器
    num_lgs = math.ceil(total_servers / servers_per_lg)
    num_sgs = num_lgs  # SG的数量通常与LG相同

    # 初始化连接列表
    connections = []

    # 生成Leaf到Spine的连接
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            for spine_index in range(spines_per_sg):
                leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
                spine_name = f"SG{lg_index+1}-Spine{spine_index+1}"
                for port_group in range(32 // spines_per_sg):
                    leaf_port = spine_index * (32 // spines_per_sg) + port_group + 1
                    spine_port = leaf_index * (32 // spines_per_sg) + port_group + 1
                    connections.append((leaf_name, f"Port{leaf_port}", spine_name, f"Port{spine_port}"))

    # 生成Spine到Core的连接
    for sg_index in range(num_sgs):
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{sg_index+1}-Spine{spine_index+1}"
            # 为每个Spine交换机确定连接到Core交换机的端口范围
            spine_ports = list(range(33 + spine_index * 4, 37 + spine_index * 4))
            for core_index in range(cores_per_cg):
                core_name = f"CG{spine_index+1}-Core{core_index+1}"
                # 为每个Core交换机确定连接到Spine交换机的端口范围
                core_ports = list(range(1 + sg_index * 4, 5 + sg_index * 4))
                for spine_port, core_port in zip(spine_ports, core_ports):
                    connections.append((spine_name, f"Port{spine_port}", core_name, f"Port{core_port}"))

    # 生成GPU到Leaf的连接
    gpus_per_lg = 32
    hcas_per_gpu = 8
    for lg_index in range(num_lgs):
        # 如果是最后一个LG，确定剩余的GPU数量
        gpus_in_this_lg = gpus_per_lg if (lg_index + 1) < num_lgs else total_servers - lg_index * gpus_per_lg
        for gpu_index in range(gpus_in_this_lg):
            for hca_index in range(hcas_per_gpu):
                su_name = f"SU{lg_index+1}"
                gpu_name = f"{su_name}-GPU{gpu_index+1}"
                hca_name = f"{gpu_name}"
                hca_port = f"HCA{hca_index+1}"
                leaf_name = f"LG{lg_index+1}-Leaf{hca_index+1}"
                leaf_port = 33 + gpu_index
                connections.append((hca_name, hca_port, leaf_name, f"Port{leaf_port}"))

    return connections

def generate_three_tier_topology_1024(total_servers, num_cgs=8, leaves_per_lg=8, spines_per_sg=8, cores_per_cg=16, ports_per_leaf=64, ports_per_spine=64):
    # Calculate the number of LGs and SGs based on the total number of servers
    servers_per_lg = 32  # Assuming each LG has 32 servers
    num_lgs = math.ceil(total_servers / servers_per_lg)
    num_sgs = num_lgs  # SGs number is usually the same as LGs

    # Initialize connection list
    connections = []

    # Generate Leaf to Spine connections
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            for spine_index in range(spines_per_sg):
                leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
                # The spine group index is based on the modulus of the LG index
                spine_name = f"SG{(lg_index % num_sgs) + 1}-Spine{spine_index+1}"
                for port_group in range(4):  # 32 ports divided into 8 groups, 4 cables per group
                    leaf_port = spine_index * 4 + port_group + 1
                    spine_port = leaf_index * 4 + port_group + 1
                    connections.append((leaf_name, f"Port{leaf_port}", spine_name, f"Port{spine_port}"))

    # Generate Spine to Core connections
    for sg_index in range(num_sgs):
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{sg_index+1}-Spine{spine_index+1}"
            spine_ports = list(range(33, 65))  # Using the second half of ports (33-64)
            for core_index in range(cores_per_cg):
                core_name = f"CG{spine_index+1}-Core{core_index+1}"
                core_ports = list(range(sg_index * 2 + 1, sg_index * 2 + 3))  # 2 cables per connection
                for spine_port, core_port in zip(spine_ports[core_index*2:(core_index+1)*2], core_ports):
                    connections.append((spine_name, f"Port{spine_port}", core_name, f"Port{core_port}"))

    # Generate GPU to Leaf connections
    gpus_per_lg = 32
    hcas_per_gpu = 8
    for lg_index in range(num_lgs):
        # Determine the number of GPUs for this LG, which could be less for the last LG
        gpus_in_this_lg = gpus_per_lg if (lg_index + 1) < num_lgs else total_servers - lg_index * gpus_per_lg
        for gpu_index in range(gpus_in_this_lg):
            for hca_index in range(hcas_per_gpu):
                su_name = f"SU{lg_index+1}"
                gpu_name = f"{su_name}-GPU{gpu_index+1}"
                hca_name = f"{gpu_name}"
                hca_port = f"HCA{hca_index+1}"
                leaf_name = f"LG{lg_index+1}-Leaf{hca_index+1}"
                leaf_port = 33 + gpu_index
                connections.append((hca_name, hca_port, leaf_name, f"Port{leaf_port}"))

    return connections

def generate_three_tier_topology_2048(total_servers, num_cgs=8, leaves_per_lg=8, spines_per_sg=8, cores_per_cg=32, ports_per_leaf=64, ports_per_spine=64):
    # Calculate the number of LGs and SGs based on the total number of servers
    servers_per_lg = 32  # Assuming each LG has 32 servers
    num_lgs = math.ceil(total_servers / servers_per_lg)
    num_sgs = num_lgs  # SGs number is usually the same as LGs

    # Initialize connection list
    connections = []

    # Generate Leaf to Spine connections
    for lg_index in range(num_lgs):
        for leaf_index in range(leaves_per_lg):
            for spine_index in range(spines_per_sg):
                leaf_name = f"LG{lg_index+1}-Leaf{leaf_index+1}"
                spine_name = f"SG{(lg_index % num_sgs)+1}-Spine{spine_index+1}"
                for port_group in range(4):  # 32 ports divided into 8 groups, 4 cables per group
                    leaf_port = spine_index * 4 + port_group + 1
                    spine_port = leaf_index * 4 + port_group + 1
                    connections.append((leaf_name, f"Port{leaf_port}", spine_name, f"Port{spine_port}"))

    # Generate Spine to Core connections
    for sg_index in range(num_sgs):
        for spine_index in range(spines_per_sg):
            spine_name = f"SG{sg_index+1}-Spine{spine_index+1}"
            spine_ports = list(range(33, 65))  # Using the second half of ports (33-64)
            for core_index in range(cores_per_cg):
                cg_index = spine_index  # CG index corresponds to Spine index
                core_name = f"CG{cg_index+1}-Core{core_index+1}"
                core_ports = list(range(sg_index + 1, sg_index + 2))  # 1 cable per connection
                for spine_port, core_port in zip(spine_ports[core_index:core_index+1], core_ports):
                    connections.append((spine_name, f"Port{spine_port}", core_name, f"Port{core_port}"))

    # Generate GPU to Leaf connections
    gpus_per_lg = 32
    hcas_per_gpu = 8
    for lg_index in range(num_lgs):
        # Determine the number of GPUs for this LG, which could be less for the last LG
        gpus_in_this_lg = min(gpus_per_lg, total_servers - lg_index * gpus_per_lg)
        for gpu_index in range(gpus_in_this_lg):
            for hca_index in range(hcas_per_gpu):
                su_name = f"SU{lg_index+1}"
                gpu_name = f"{su_name}-GPU{gpu_index+1}"
                hca_name = f"{gpu_name}"
                hca_port = f"HCA{hca_index+1}"
                leaf_name = f"LG{lg_index+1}-Leaf{hca_index+1}"
                leaf_port = 33 + gpu_index
                connections.append((hca_name, hca_port, leaf_name, f"Port{leaf_port}"))

    return connections

def generate_topology(total_servers):
    if 0 <= total_servers <= 256:
        return generate_flexible_topology_256(total_servers)
    elif 257 <= total_servers <= 512:
        return generate_three_tier_topology_512(total_servers)
    elif 513 <= total_servers <= 1024:
        return generate_three_tier_topology_1024(total_servers)
    elif 1025<= total_servers <= 2048:
        return generate_three_tier_topology_2048(total_servers)
    else:
        print("Only the range between 1 and 2048 is supported.")
        exit()


def write_to_excel(connections, filename):
    # 创建一个新的工作簿
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "GPU to Leaf"
    ws2 = wb.create_sheet(title="Leaf to Spine")
    ws3 = wb.create_sheet(title="Spine to Core")

    # 写入表头
    for ws in (ws1, ws2, ws3):
        ws.append(['Source Device', 'Source Port', 'Destination Device', 'Destination Port'])

    # 根据连接类型将数据写入相应的工作表
    for conn in connections:
        if "GPU" in conn[0] and "Leaf" in conn[2]:
            ws1.append(conn)
        elif "Leaf" in conn[0] and "Spine" in conn[2]:
            ws2.append(conn)
        elif "Spine" in conn[0] and "Core" in conn[2]:
            ws3.append(conn)

    # 保存工作簿
    wb.save(filename)

# 请确保 main 函数中调用了 write_to_excel 而不是 write_to_csv，并且传递了正确的文件扩展名
def main():
    total_servers = int(input("Enter the total number of servers: "))
    connections = generate_topology(total_servers)
    
    # Print connections to screen
    print("Connections:")
    for conn in connections:
        print(f"{conn[0]}:{conn[1]} -> {conn[2]}:{conn[3]}")
    
    # Write connections to Excel file
    excel_filename = f"topology_{total_servers}_servers.xlsx"
    write_to_excel(connections, excel_filename)
    print(f"\nConnections have been written to {excel_filename}")

if __name__ == "__main__":
    main()

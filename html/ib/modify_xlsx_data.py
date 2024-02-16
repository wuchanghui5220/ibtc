#!/usr/bin/python3

import openpyxl
import re
import sys


rules_file = sys.argv[1]
data_file = sys.argv[2]

# 打开replace_rules.xlsx文件
workbook = openpyxl.load_workbook(rules_file)
sheet = workbook.active

# 构建替换规则字典
replace_rules = {}
for row in range(2, sheet.max_row + 1):  # 从第二行开始读取，跳过标题行
    template = sheet[f'A{row}'].value  # 读取模板名称
    project = sheet[f'B{row}'].value  # 读取项目名称
    if template and project:  # 确保模板名称和项目名称不为空
        replace_rules[template] = project

# 现在replace_rules字典包含了从Excel文件中读取的数据
# print(replace_rules)

# 接下来，您可以使用这个字典在您的其他脚本中进行替换操作

# 打开Excel文件
workbook = openpyxl.load_workbook(data_file)

# 遍历所有工作表并应用替换规则
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # 遍历工作表中的所有行和列
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=False):
        for cell in row:
            if isinstance(cell.value, str):  # 检查是否为字符串
                for key, value in replace_rules.items():
                    # 匹配字符串开头的关键字
                    if re.match(r'^' + re.escape(key) + r'(\b|_)', cell.value):
                        cell.value = re.sub(r'^' + re.escape(key), value, cell.value)
                        break  # 匹配到一个就替换，然后跳出循环

# 保存到新的Excel文件
workbook.save(f'modified_{data_file}')

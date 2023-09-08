import pandas as pd
import openpyxl
import sys
import glob


files = {
    # 'HW': [],
    'SWS X': [],
    'SWS P': [] 
}

# hw_files = glob.glob('*HW*.xlsx')
# for f in hw_files:
#     files['HW'].append(f)

sws_x_files = glob.glob('*SWS X*.xlsx')  
for f in sws_x_files:
    files['SWS X'].append(f)
    
sws_p_files = glob.glob('*SWS P*.xlsx')
for f in sws_p_files:
    files['SWS P'].append(f)
    
print(files)
file1 = sys.argv[1]
file2 = files['SWS X'][0]
file3 = files['SWS P'][0]

# 读取第一张表格  
hw_df = pd.read_excel(file1, header=None)

# 设置索引,转换为字典
search_criteria = hw_df.set_index(hw_df.index).iloc[:,3].to_dict()

# 按原顺序遍历结果字典  
for key in search_criteria:
    print(search_criteria[key])

# 读取第二张表格
sws_x_df = pd.read_excel(file2, skiprows=3)

# 遍历搜索条件字典,逐行搜索匹配    
sws_x_results = {}
for key in search_criteria:
    crit = search_criteria[key]
    res = sws_x_df[sws_x_df.iloc[:,0]==crit]
    if not res.empty:
        sws_x_results[key] = res.iloc[0, 2]

# 按原顺序遍历结果字典  
for key in sws_x_results:
    print(sws_x_results[key])

# 读取第三张表格        
sws_df = pd.read_excel(file3, skiprows=1)

# 遍历第二张表结果,在第三张表搜索       
N_cols = {}
Q_cols = {}
for key in sws_x_results:
    part = sws_x_results[key]
    res = sws_df[sws_df.iloc[:,17]==part]
    if not res.empty:
        N_cols[key] = res.iloc[0, 13]
        Q_cols[key] = res.iloc[0, 16]

# 按原顺序遍历结果字典  
for key in N_cols:
    print(N_cols[key])
    
for key in Q_cols:
    print(Q_cols[key])

# wb = openpyxl.Workbook()
# sheet = wb.active

# # 写入第一列  

# for i in range(len(search_criteria)):
#     sheet.cell(i+1, 1).value = search_criteria[i]

# # 写入第二列
# for i in range(len(sws_x_results)): 
#     sheet.cell(i+1, 2).value = sws_x_results[i]

# # 写入第三列
# for i in range(len(N_cols)):
#     sheet.cell(i+1, 3).value = N_cols[i]

# # 写入第四列  
# for i in range(len(Q_cols)):
#     sheet.cell(i+1, 4).value = Q_cols[i]
    
# wb.save('sup.xlsx')

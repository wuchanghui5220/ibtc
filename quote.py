import pandas as pd
import openpyxl
import sys
import os
import time
import re
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

def is_chinese(value):
  value = str(value)

  for ch in value: 
    if u'\u4e00' <= ch <= u'\u9fff':
      return True

  return False

excel_file = sys.argv[1]

if len(sys.argv) < 2:
    print("Usage: python demo.py <excel_file>")
    sys.exit(1)

if not os.path.exists(sys.argv[1]):
    print("Excel file does not exist")
    sys.exit(1)  
# df = pd.read_excel(excel_file, sheet_name='Q423_NBU_HW') 
df = pd.read_excel(excel_file) 
df = df.loc[:,['Price Book Organizer', 'Product Name','Product Description','P3 = OEM Price']]

数量_dict = {} 

def search_and_select(df):
    while True:
        search_term = input("请输入搜索词:")
        search_term = search_term.upper()
        search_result = df[df['Product Name'].str.contains(search_term)]

        if search_result.empty:
            print("未找到结果,请重新输入!")
        else:
            break
    
    max_len = len(max(search_result['Product Name'], key=len))
    
    for i, row in enumerate(search_result[['Product Name','Product Description']].iterrows()):
        if i <= 9:
            name = row[1]['Product Name'].ljust(max_len + 1)  
            desc = row[1]['Product Description']
            print(f"{i}. {name} {desc}  [{i}]")
        elif i > 9:
            name = row[1]['Product Name'].ljust(max_len)  
            desc = row[1]['Product Description']
            print(f"{i}. {name} {desc}  [{i}]")

    need_PN = []   
    while True:
        try:
            choice = input("请输入要选择的序号(直接回车结束本次选择):")
            if choice == '':
                break
            
            choice = int(choice)
            if choice >= 0 and choice < len(search_result):
                selected = search_result.iloc[choice]['Product Name']
                need_PN.append(selected)
                # 提示输入数量
                数量 = int(input("请输入{}的购买数量:".format(selected)))
                # 数量_dict[selected] = 数量 
                if selected in 数量_dict:
                    数量_dict[selected] += 数量
                else:
                    数量_dict[selected] = 数量
            else:
                print("没有这个选项,请重新输入")
                
        except ValueError:
            print("请输入一个数字!")
        except KeyboardInterrupt:
            break
    
    print("本次选择产品:")      
    print(need_PN)
    print()
    
    return need_PN

def get_ex_rate(prompt, default=7.35):
    rate = input(prompt) or default
    try:
        return float(rate)
    except ValueError:
        print("输入无效, 使用默认值")
        return default
def get_npn_rate(prompt, default=10):
    rate = int(input(prompt) or default)
    rate = float(rate / 100)
    try:
        return rate
    except ValueError:
        print("输入无效, 使用默认值")
        return default

# Welcome!
print("\n欢迎使用快速报价小程序，输入产品 PN 快速搜索！\n")
selected_PN = []
while True:
    pn = search_and_select(df)
    selected_PN.extend(pn)
    
    again = input("是否需要再次搜索(Y/N)?")
    if again == "n" or again == "N":
        break

df = df[df['Product Name'].isin(selected_PN)]
current_time = time.strftime("%Y%m%d-%H%M%S")
xlsx_file_name = f'NVIDIA-Networking-Product-Quote-正阳恒卓-{current_time}.xlsx'
xlsx_file_name_new = f'NVIDIA-Networking-Product-Quote-正阳恒卓报价单-{current_time}.xlsx'
# df.to_excel(xlsx_file_name, index=False)

# 在'P3 = OEM Price'后插入新列'汇率'
ER = get_ex_rate("输入人民币对美元汇率(默认值7.35): ")
df.insert(loc=4, column='汇率', value=ER) 

# 在'汇率'后插入新列'税费'
df.insert(loc=5, column='税费', value=0.13)

# 在'税费'后插入新列'总代利润点' 
df.insert(loc=6, column='总代利润点', value=0.05)

# 在'总代利润点'后插入新列'NPN利润点'
NR = get_npn_rate("输入利润点(默认 10 个点): ")
df.insert(loc=7, column='NPN利润点', value=NR)

# 新增一列'数量'
df['数量'] = None
for product, 数量 in 数量_dict.items():
    df.loc[df['Product Name'] == product, '数量'] = 数量

# 在'数量'后插入新列'含税单价', 计算含税单价
df.insert(loc=9, column='含税单价', value=df['P3 = OEM Price'] * df['汇率'] * (1 + df['税费']) * (1 + df['总代利润点']) * (1 + df['NPN利润点']))

# 在'含税单价'后插入新列'含税总价', 计算含税总价
df.insert(loc=10, column='含税总价', value=df['数量'] * df['含税单价'])

# 重新格式化含税单价为整数
df['含税单价'] = df['含税单价'].astype(int)

# 重新格式化含税总价为整数  
df['含税总价'] =df['含税总价'].astype(int)

# 数据写入到文件，再加载到openpyxl，进行二次编辑
# df.to_excel(xlsx_file_name, index=False)
# wb = openpyxl.load_workbook(xlsx_file_name)
# ws = wb.worksheets[0]

# 数据不写入到文件，直接append到wb的ws，进行二次编辑
wb = openpyxl.Workbook()
ws = wb.active
for index, row in df.iterrows():
    ws.append(row.tolist())

# 设置D列为美元格式 
for i in range(1, ws.max_row+1):
  c_col = 'D' + str(i)
  ws[c_col].number_format = u'$#,##0.00' 



# 设置 F，G, H 列为百分比 % 格式
for i in range(1, ws.max_row+1):
  e_col = 'F' + str(i)
  f_col = 'G' + str(i) 
  g_col = 'H' + str(i)
  
  ws[e_col].number_format = '0%'
  ws[f_col].number_format = '0%'
  ws[g_col].number_format = '0%'

# 设置 '数量' 这一列 I 列 为整数
for i in range(1, ws.max_row+1):
  c_col = 'I' + str(i)
  ws[c_col].number_format = '0' 

# 设置J，K 列为货币RMB格式
for i in range(1, ws.max_row+1):
    i_col = 'J' + str(i) 
    j_col = 'K' + str(i)
    ws[i_col].number_format = u'¥#,##0.00'
    ws[j_col].number_format = u'¥#,##0.00' 


# 在第9列(I列)前插入一列
ws.insert_cols(9)  

# 在新插入的这一列写入值
for cell in ws["I"]:
    cell.value = "NVIDIA"

# 在第一列前插入新列
ws.insert_cols(1)

# 插入序号
for i in range(1, ws.max_row+1):
    c = ws.cell(i, 1) 
    c.value = i
    c.number_format = '0' # 指定数字格式为General "常规"

# 获取最后一列的索引 
last_col = ws.max_column 

# 在最后一列后插入新列
ws.insert_cols(last_col+1) 

# 在新插入的这一列写入值
for cell in ws["N"]:
    cell.value = "8-20周"

# 设置英文字体
# font_en = openpyxl.styles.Font(name='Tahoma', size=11)

# # 设置中文字体 
# font_cn = openpyxl.styles.Font(name='微软雅黑', size=11)

# for row in ws.rows:
#     for cell in row:
#         if is_chinese(cell.value):
#             cell.font = font_cn
#         else:
#             cell.font = font_en

# 获取表格的所有单元格
last_col = get_column_letter(ws.max_column) 
cells = ws[f'A1:{last_col}{ws.max_row}']
# 表格的所有单元格垂直居中，水平居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center')

# 要实现左对齐的列号, 开启自动换行
columns = ['B', 'D']

for column in columns:
    for cell in ws[column]:
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

# 在第一列前插入新列
ws.insert_cols(1)
# 添加边框
thin = Side(border_style="thin", color="000000")

for row in ws.rows:
    for cell in row:
        if cell.value:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

# 设置列宽
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 21
ws.column_dimensions['D'].width = 21
ws.column_dimensions['E'].width = 81
ws.column_dimensions['F'].width = 21
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 18
ws.column_dimensions['N'].width = 18
ws.column_dimensions['O'].width = 10

wb.save(xlsx_file_name_new)

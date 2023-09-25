import pandas as pd
import openpyxl
import sys
import os
import time
import random
import re
import glob
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from termcolor import colored  

def is_chinese(value):
  value = str(value)

  for ch in value: 
    if u'\u4e00' <= ch <= u'\u9fff':
      return True

  return False


files = {
    'HW': [],
    'SWS X': [],
    'SWS P': [] 
}
hw_files = glob.glob('*HW*.xlsx')  
for f in hw_files:
    files['HW'].append(f)
sws_x_files = glob.glob('*SWS X*.xlsx')  
for f in sws_x_files:
    files['SWS X'].append(f)
    
sws_p_files = glob.glob('*SWS P*.xlsx')
for f in sws_p_files:
    files['SWS P'].append(f)
    
# print(files)
file1 = files['HW'][0]
file2 = files['SWS X'][0]
file3 = files['SWS P'][0]

template_file = 'template.xlsx'

# if len(sys.argv) < 2:
#     print("Usage: python demo.py <excel_file>")
#     sys.exit(1)

if not os.path.exists(file1):
    print("Excel file does not exist")
    sys.exit(1)  
# df = pd.read_excel(excel_file, sheet_name='Q423_NBU_HW') 
df1 = pd.read_excel(file1) 
df1 = df1.loc[:,['Price Book Organizer', 'Product Name','Product Description','P3 = OEM Price']]

# product qty dict
qty_dict = {} 
# license qty dict
lic_qty_dict = {}
# Product service life
PSL_list = [1, 2, 3, 4, 5]
PSL_num = 0
all_psl = False
def search_and_select(df):
    last_color = None
    colors = ['white', 'green','yellow','blue','magenta','cyan']
    while True:
        search_term = input("请输入搜索词:")
        search_term = search_term.upper()
        search_result = df1[df1['Product Name'].str.contains(search_term)]

        if search_result.empty:
            print("未找到结果,请重新输入!")
        else:
            break
    
    max_len = len(max(search_result['Product Name'], key=len))
    
    for i, row in enumerate(search_result[['Product Name','Product Description', 'P3 = OEM Price']].iterrows()):
        while True:
            c = random.choice(colors)
            if c != last_color:
                break
        last_color = c
        if i <= 9:
            name = row[1]['Product Name'].ljust(max_len + 1)  
            desc = row[1]['Product Description']
            price = row[1]['P3 = OEM Price']
            print(colored(f"{i}. {name} {desc} ${price} ", c))
        elif i > 9:
            name = row[1]['Product Name'].ljust(max_len)  
            desc = row[1]['Product Description']
            price = row[1]['P3 = OEM Price']
            # print(f"{i}. {name} {desc} ${price} [{i}]")
            print(colored(f"{i}. {name} {desc} ${price} ", c))

    need_PN = []   
    while True:
        try:
            choice = input("请输入要选择的序号(直接回车结束本次选择):")
            if choice == '':
                break
            
            choice = int(choice)
            if choice >= 0 and choice < len(search_result):
                selected = search_result.iloc[choice]['Product Name']
                need_PN.append(selected)
                # 提示输入数量
                qty = int(input("请输入{}的购买数量:".format(selected)))
                # 数量_dict[selected] = 数量 
                if selected in qty_dict:
                    qty_dict[selected] += qty
                else:
                    qty_dict[selected] = qty

            else:
                print("没有这个选项,请重新输入")
                
        except ValueError:
            print("请输入一个数字!")
        except KeyboardInterrupt:
            break
    
    print("已选择产品和数量:")      
    print(qty_dict)
    print()
    
    return need_PN

dfx = pd.read_excel(file2, skiprows=3) 
dfx = dfx.loc[:,['Configured SKU: Product Name', 'Configured SKU: Product Code','Product Name','Product Code']]
def search_and_select_sws_x(df):
    while True:
        search_term = input("请输入搜索词:")
        search_term = search_term.upper()
        search_result = dfx[dfx['Configured SKU: Product Name'].str.contains(search_term, na=False)]

        if search_result.empty:
            print("未找到结果,请重新输入!")
        else:
            break
    
    max_len = len(max(search_result['Configured SKU: Product Name'], key=len))
    
    for i, row in enumerate(search_result[['Configured SKU: Product Name','Product Name']].iterrows()):
        if i <= 9:
            name = row[1]['Configured SKU: Product Name'].ljust(max_len + 1)  
            desc = row[1]['Product Name']
            print(f"{i}. {name} {desc}  [{i}]")
        elif i > 9:
            name = row[1]['Configured SKU: Product Name'].ljust(max_len)  
            desc = row[1]['Product Name']
            print(f"{i}. {name} {desc}  [{i}]")

    need_PN = []   
    while True:
        try:
            choice = input("请输入要选择的序号(直接回车结束本次选择):")
            if choice == '':
                break
            
            choice = int(choice)
            if choice >= 0 and choice < len(search_result):
                selectedx = search_result.iloc[choice]['Configured SKU: Product Name']
                need_PN.append(selectedx)
                # 提示输入数量
                qtyx = int(input("请输入{}的购买数量:".format(selectedx)))
                 
                if selectedx in qty_dict:
                    qty_dict[selectedx] += qtyx
                else:
                    qty_dict[selectedx] = qtyx

            else:
                print("没有这个选项,请重新输入")
                
        except ValueError:
            print("请输入一个数字!")
        except KeyboardInterrupt:
            break
    
    print("已选择的产品服务和数量:")      
    print(need_PN, qtyx)
    print()
    
    return need_PN

dfsws = pd.read_excel(file3) 
dfsws = dfsws.loc[:,['Price Book Organizer', 'Mellanox Legacy', 'Description','P8 - NPN PRICE']]
def search_and_select_sws(df):
    while True:
        search_term = input("请输入搜索词:")
        search_term = search_term.upper()
        search_result = dfsws[dfsws['Mellanox Legacy'].str.contains(search_term, na=False)]

        if search_result.empty:
            print("未找到结果,请重新输入!")
        else:
            break
    
    max_len = len(max(search_result['Mellanox Legacy'], key=len))
    
    for i, row in enumerate(search_result[['Mellanox Legacy','Description']].iterrows()):
        if i <= 9:
            name = row[1]['Mellanox Legacy'].ljust(max_len + 1)  
            desc = row[1]['Description']
            print(f"{i}. {name} {desc}  [{i}]")
        elif i > 9:
            name = row[1]['Mellanox Legacy'].ljust(max_len)  
            desc = row[1]['Description']
            print(f"{i}. {name} {desc}  [{i}]")

    need_PN = []   
    while True:
        try:
            choice = input("请输入要选择的序号(直接回车结束本次选择):")
            if choice == '':
                break
            
            choice = int(choice)
            if choice >= 0 and choice < len(search_result):
                selectedsws = search_result.iloc[choice]['Mellanox Legacy']
                need_PN.append(selectedsws)
                # 提示输入数量
                qtysws = int(input("请输入{}的购买数量:".format(selectedsws)))
                # 数量_dict[selected] = 数量 
                if selectedsws in lic_qty_dict:
                    lic_qty_dict[selectedsws] += qtysws
                else:
                    lic_qty_dict[selectedsws] = qtysws

            else:
                print("没有这个选项,请重新输入")
                
        except ValueError:
            print("请输入一个数字!")
        except KeyboardInterrupt:
            break
    
    print("已选择的license和数量:")      
    print(need_PN, qtysws)
    print()
    
    return need_PN

def get_ex_rate(prompt, default=7.40):
    rate = input(prompt) or default
    try:
        return float(rate)
    except ValueError:
        print("输入无效, 使用默认值")
        return default
def get_npn_rate(prompt, default=10):
    rate = int(input(prompt) or default)
    rate = float(rate / 100)
    try:
        return rate
    except ValueError:
        print("输入无效, 使用默认值")
        return default

# Welcome!
# print("\n欢迎使用快速报价小程序，输入产品 PN 快速搜索！\n")
selected_PN = []
while True:
    pn = search_and_select(df1)
    selected_PN.extend(pn)
    
    again = input("是否需要再次搜索(默认y/Y)?")
    if again == "n" or again == "N":
        while True:
            PSL_num_input = input("请输入交换机银牌服务的年限,默认1年,范围[1~5]年: ")
            if PSL_num_input == "":
                PSL_num = 1
                print("使用默认值：1\n")
                break
            elif PSL_num_input == "1":
                PSL_num = 1
                print("已选择：1\n")
                break
            elif PSL_num_input.isdigit():
                PSL_num = int(PSL_num_input)
                if PSL_num in PSL_list:
                    break
                else:
                    print("输入年限不在范围内,请重新输入!")
            else:
                print("输入有误,请重新输入数字:")
        all_psl_input = input("非交换机产品的服务是否一并购买（默认n/N): ")
        if all_psl_input == 'y' or all_psl_input == 'Y' or all_psl_input == 'yes':
            all_psl = True
        elif all_psl_input == '' or all_psl_input == 'n' or all_psl_input == 'N':
            print("已选择不购买非交换机产品的服务！\n")
        break

df1 = df1[df1['Product Name'].isin(selected_PN)]
current_time = time.strftime("%Y%m%d-%H%M%S")
xlsx_file_name = f'正阳恒卓-{current_time}.xlsx'
xlsx_file_name_new = f'正阳恒卓报价单-留存单-{current_time}.xlsx'

# 获取df1的表头
# print("df1 columns:", df1.columns)
df1_columns = df1.columns
# print("df1_columns:", df1_columns)

# print("产品和数量字典：")
# print(qty_dict)
# -----------------------------------------------------------------------------------


# 将PN列表转换为DataFrame
# 将产品名称列表直接作为搜索条件字典
# search_criteria = {pn: pn for pn in selected_PN}

# print(search_criteria)

# # 读取第二张表格
# sws_x_df = pd.read_excel(file2, skiprows=3)

# # 遍历搜索条件字典,逐行搜索匹配    
# sws_x_results = {}
# for key in search_criteria:
#     crit = search_criteria[key]
#     res = sws_x_df[sws_x_df.iloc[:,0]==crit]
#     if not res.empty:
#         sws_x_results[key] = res.iloc[0, 2]

# print(sws_x_results)


#------------------------------
license_PN = []
buy_lic = input('''是否购买UFM或者RiverMAX软件license？
    [默认n/N] 
    [输入 l/L 查询license]
    请输入你的选择： ''')
# if buy_lic == 'R' or buy_lic == 'r':
#     pn = search_and_select_sws_x(dfx)
#     license_PN.extend(pn)
if buy_lic == 'L' or buy_lic == 'l':
    pn = search_and_select_sws(dfsws)
    license_PN.extend(pn)
elif buy_lic == '':
    print("用户选择不需要购买license！\n")
    pass
else:
    print("用户选择不需要购买license！\n")
    pass

search_criteria = {pn: pn for pn in selected_PN} 
# 按原顺序遍历结果字典  
# print("第一张表格的筛选结果：")
# print(search_criteria)
# for key in search_criteria:
#     print(search_criteria[key])

# 读取第二张表格，sws x表格数据
sws_x_df = pd.read_excel(file2, skiprows=3)

# 遍历搜索条件字典,逐行搜索匹配    
sws_x_results = {}
for key in search_criteria:
    crit = search_criteria[key]
    res = sws_x_df[sws_x_df.iloc[:,0]==crit]
    if not res.empty:
        sws_x_results[key] = res.iloc[0, 2]

# 按原顺序遍历结果字典 
# print("第二张表格搜索结果") 
# print(sws_x_results)
# 修改sws_x_results 服务1G为nS
for key, value in sws_x_results.items():
    if '1G' in value:
        new_value = value.replace('1G', str(PSL_num) + 'S')
        # print(f"将 {key} 的值从 {value} 修改为 {new_value}")
        sws_x_results[key] = new_value

    elif all_psl:
        if PSL_num in PSL_list[1:] and '2B' in value:
            new_value = value.replace('2B', str(PSL_num) + 'B')
            sws_x_results[key] = new_value
# print("第二张表格修改服务的结果") 
# print(sws_x_results)

#------------------------------
# 服务和数量字典
service_qty_dict = {}
for product, qty in qty_dict.items():
    service = sws_x_results[product]
    if service in service_qty_dict:
        service_qty_dict[service] += qty
    else:
        service_qty_dict[service] = qty
# print("第二张表格服务和数量的结果")     
# print(service_qty_dict)
#---------------------------------------
# 使用 set 存储已经见过的 value
seen_values = set() 
# 遍历并删除重复值,保留第一个
for key, value in list(sws_x_results.items()):
    if value in seen_values:
        del sws_x_results[key]
    else:
        seen_values.add(value)
# print("第二张表格删除重复维保名称的服务和数量的结果")  
# print(sws_x_results)

# 删除非交换机设备的维保服务
if not all_psl:
    for key in list(sws_x_results):
        if 'SUP' not in sws_x_results[key]:
            del sws_x_results[key]
# print("第二张表格删除用户选择的不需要的名称的服务和数量的结果") 
# print(sws_x_results)
# print("第二张表格操作结束！")
#-----------------------------------------------
# print("打印已选择的PN")
# print(qty_dict)
# print(service_qty_dict)
# print(lic_qty_dict)
# 读取第三张表格        
sws_df = pd.read_excel(file3, skiprows=1)

# 遍历第二张表结果,在第三张表搜索 

A_cols = {}    
N_cols = {}
Q_cols = {}

# 把license key/value 添加到sws_x_results
license_dict = {pn: pn for pn in license_PN}
for k, v in license_dict.items():
    if k not in sws_x_results:
        sws_x_results[k] = v
# print("查看添加license key/value后的结果")
# print(sws_x_results)
for key in sws_x_results:
    part = sws_x_results[key]
    res = sws_df[sws_df.iloc[:,17]==part]
    if not res.empty:
        A_cols[key] = res.iloc[0, 1]
        N_cols[key] = res.iloc[0, 13]
        Q_cols[key] = res.iloc[0, 16]


# 初始化df2
df2 = pd.DataFrame()

for key in sws_x_results:

    # 构造temp_df,并设置列名为df1前4列  
    temp_df = pd.DataFrame(columns=df1_columns[:4], data={
      df1_columns[0]: A_cols[key], 
      df1_columns[1]: [sws_x_results[key]],
      df1_columns[2]: [N_cols[key]],
      df1_columns[3]: [Q_cols[key]]
    })

    # 获取temp_df列名
    cols = list(temp_df.columns)

    # 使用cols子集执行dropna

    subset = temp_df.columns.tolist() 
    # print(subset)
    temp_df = temp_df.dropna(axis=1, how='all') 
    # print(subset) 

    # 拼接到df2
    df2 = pd.concat([df2, temp_df])

# 重置索引  
df2 = df2.reset_index(drop=True)

# 拼接到df
df = pd.concat([df1, df2])
# ------------------------------------------------------------------------------------

# df.to_excel(xlsx_file_name, index=False)

# 在'P3 = OEM Price'后插入新列'汇率'
ER = get_ex_rate("输入人民币对美元汇率(默认值7.40): ")
df.insert(loc=4, column='汇率', value=ER) 

# 在'汇率'后插入新列'税费'
df.insert(loc=5, column='税费', value=0.13)

# 在'税费'后插入新列'总代利润点' 
df.insert(loc=6, column='总代利润点', value=0.05)

# 在'总代利润点'后插入新列'NPN利润点'
NR = get_npn_rate("输入利润点(默认 10 个点): ")
df.insert(loc=7, column='NPN利润点', value=NR)

# 新增一列'数量'
df['数量'] = None
for product, qty in qty_dict.items():
    df.loc[df['Product Name'] == product, '数量'] = qty
for service, qty in service_qty_dict.items():
    df.loc[df['Product Name'] == service, '数量'] = qty
# 添加license 数量
for license, qty in lic_qty_dict.items():
    df.loc[df['Product Name'] == license, '数量'] = qty

# 在'数量'后插入新列'含税单价', 计算含税单价
df.insert(loc=9, column='含税单价', value=(df['P3 = OEM Price'] * df['汇率'] * (1 + df['税费']) * (1 + df['总代利润点']) * (1 + df['NPN利润点'])).astype(int))

# 在'含税单价'后插入新列'含税总价', 计算含税总价
df.insert(loc=10, column='含税总价', value=df['数量'] * df['含税单价'])

# 重新格式化含税单价为整数
# df['含税单价'] = df['含税单价'].astype(int)

# 重新格式化含税总价为整数  
# df['含税总价'] =df['含税总价'].astype(int)

# 数据写入到文件，再加载到openpyxl，进行二次编辑
# df.to_excel(xlsx_file_name, index=False)
# wb = openpyxl.load_workbook(xlsx_file_name)
# ws = wb.worksheets[0]

# 数据不写入到文件，直接append到wb的ws，进行二次编辑
wb = openpyxl.Workbook()
ws = wb.active
for index, row in df.iterrows():
    ws.append(row.tolist())

# 设置D列为美元格式 
for i in range(1, ws.max_row+1):
  c_col = 'D' + str(i)
  ws[c_col].number_format = u'$#,##0.00' 



# 设置 F，G, H 列为百分比 % 格式
for i in range(1, ws.max_row+1):
  e_col = 'F' + str(i)
  f_col = 'G' + str(i) 
  g_col = 'H' + str(i)
  
  ws[e_col].number_format = '0%'
  ws[f_col].number_format = '0%'
  ws[g_col].number_format = '0%'

# 设置 '数量' 这一列 I 列 为整数
for i in range(1, ws.max_row+1):
  c_col = 'I' + str(i)
  ws[c_col].number_format = '0' 

# 设置J，K 列为货币RMB格式
for i in range(1, ws.max_row+1):
    i_col = 'J' + str(i) 
    j_col = 'K' + str(i)
    ws[i_col].number_format = u'¥#,##0.00'
    ws[j_col].number_format = u'¥#,##0.00' 


# 在第9列(I列)前插入一列
ws.insert_cols(9)  

# 在新插入的这一列写入值
for cell in ws["I"]:
    cell.value = "NVIDIA"

# 在第一列前插入新列
ws.insert_cols(1)

# 插入序号
for i in range(1, ws.max_row+1):
    c = ws.cell(i, 1) 
    c.value = i
    c.number_format = '0' # 指定数字格式为General "常规"

# 获取最后一列的索引 
last_col = ws.max_column 

# 在最后一列后插入新列
ws.insert_cols(last_col+1) 

# 在新插入的这一列写入值
for cell in ws["N"]:
    cell.value = "8-20周"

# 设置英文字体
# font_en = openpyxl.styles.Font(name='Tahoma', size=11)

# # 设置中文字体 
# font_cn = openpyxl.styles.Font(name='微软雅黑', size=11)

# for row in ws.rows:
#     for cell in row:
#         if is_chinese(cell.value):
#             cell.font = font_cn
#         else:
#             cell.font = font_en

# 获取表格的所有单元格
last_col = get_column_letter(ws.max_column) 
cells = ws[f'A1:{last_col}{ws.max_row}']
# 表格的所有单元格垂直居中，水平居中
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center')

# 要实现左对齐的列号, 开启自动换行
columns = ['B', 'D']

for column in columns:
    for cell in ws[column]:
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

# 在第一列前插入新列
ws.insert_cols(1)
# 添加边框
thin = Side(border_style="thin", color="000000")

for row in ws.rows:
    for cell in row:
        if cell.value:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

# 设置列宽
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 21
ws.column_dimensions['D'].width = 21
ws.column_dimensions['E'].width = 81
ws.column_dimensions['F'].width = 21
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 18
ws.column_dimensions['N'].width = 18
ws.column_dimensions['O'].width = 10

wb.save(xlsx_file_name_new)

# 打开源文件和目标文件
source_workbook = openpyxl.load_workbook(xlsx_file_name_new)
target_workbook = openpyxl.load_workbook(template_file)

# 选择源文件和目标文件中的工作表
source_sheet = source_workbook.active
target_sheet = target_workbook.active

# 选择要复制的源数据范围（B、C、D、E列，假设数据在B列中是连续的）
source_columns = ['B', 'C', 'D', 'E', 'K', 'L', 'M', 'N', 'O']
source_max_row = source_sheet.max_row

# 选择目标文件的要粘贴的位置（B16开始）
target_start_row = 16
target_start_column = 'B'

# 复制源文件的数据到目标文件
for source_column in source_columns:
    for row_index in range(1, source_max_row + 1):
        source_cell = source_sheet[source_column + str(row_index)].value
        target_row = row_index + target_start_row - 1
        target_column = chr(ord(target_start_column) + source_columns.index(source_column))
        target_sheet[target_column + str(target_row)] = source_cell


# 另存为新文件
new_filename = f'正阳恒卓报价单-{current_time}.xlsx'
target_workbook.save(new_filename)

# 关闭工作簿
source_workbook.close()
target_workbook.close()


# 打开new_filename文件
workbook = openpyxl.load_workbook(new_filename)

# 选择要操作的工作表
sheet = workbook.active

# 修改J10单元格的内容
current_date = time.strftime("%Y-%m-%d")
sheet['J10'] = f"报价日期：{current_date}"  # 将"新的内容"替换为你想要设置的新内容

# 保存修改后的文件
workbook.save(new_filename)

# 关闭工作簿
workbook.close()

print(f"报价完成，数据已保存到 [{new_filename}] 文件中, 请及时查看和检查！\n")
